from __future__ import annotations

import io
import json
import subprocess
from typing import Any

import pytest
from openpyxl import Workbook

from app.documents.indexing import IndexStagingRequest
from app.indexing.image_vlm import (
    BailianImageDescriber,
    ImageDescriptionResult,
    InternVLImageDescriber,
    NoneImageDescriber,
    image_dimensions,
    is_decorative,
)
from app.indexing.mineru import (
    MinerUAdapter,
    classify_structure,
)
from app.indexing.processing import ContentProcessor, _split_rows, merge_continuation_tables
from app.platform.config import load_platform_settings
from app.platform.errors import PlatformError


def _request(**overrides: Any) -> IndexStagingRequest:
    values: dict[str, Any] = {
        "job_id": "job_1",
        "attempt_id": "attempt_1",
        "fencing_token": 1,
        "publication_id": "publication_1",
        "document_id": "document_1",
        "document_version_id": "version_1",
        "space_id": "space_1",
        "operation": "initial",
        "base_active_version_id": None,
        "expected_generation_id": "generation_1",
        "index_revision_at_start": 0,
        "object_manifest_ref": "manifest_1",
        "processing_config_snapshot": {},
        "authorization_fence": {"actor_id": "user_1"},
        "input_manifest_hash": "manifest_hash_1",
        "processing_profile_version": "profile_1",
    }
    values.update(overrides)
    return IndexStagingRequest(**values)


class _FakeMinerU:
    def __init__(self, facts: dict[str, Any]) -> None:
        self.facts = facts
        self.calls: list[dict[str, Any]] = []

    def parse(
        self,
        content: bytes,
        *,
        media_kind: str = "",
        page_count: int = 0,
        source_url: str | None = None,
    ) -> dict[str, Any]:
        self.calls.append(
            {
                "content": content,
                "media_kind": media_kind,
                "page_count": page_count,
                "source_url": source_url,
            }
        )
        return dict(self.facts)


def _pdf_facts(**overrides: Any) -> dict[str, Any]:
    values: dict[str, Any] = {
        "text": "# Chapter\nbody",
        "page_count": 2,
        "has_text_layer": True,
        "chunks": [{"page": 1, "span": "0:5"}, {"page": 2, "span": "0:6"}],
        "structure": {"class": "tree"},
        "backend": "pipeline",
    }
    values.update(overrides)
    return values


def test_a1_routes_documents_images_and_urls_to_mineru_pipeline() -> None:
    mineru = _FakeMinerU(_pdf_facts())
    processor = ContentProcessor(mineru=mineru)
    for media_kind in (
        "application/pdf",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/vnd.openxmlformats-officedocument.presentationml.presentation",
        "text/html",
    ):
        output = processor.process(
            _request(),
            b"content",
            media_kind=media_kind,
            content_manifest_id="manifest_1",
            content_manifest_hash="manifest_hash_1",
            page_count=2,
        )
        assert output.receipt.processing_summary["route"]["mineru"] is True
        assert output.receipt.processing_summary["route"]["adapter"] == "mineru-pipeline"
    assert len(mineru.calls) == 4


def test_a1_existing_loaders_and_a4_excel_csv_never_call_mineru_or_vlm() -> None:
    def forbidden(*args: Any, **kwargs: Any) -> dict[str, Any]:
        raise AssertionError("MinerU/VLM must not be called")

    processor = ContentProcessor(mineru=forbidden, image_describer=forbidden)
    for media_kind, content in (
        ("text/plain", "plain"),
        ("text/markdown", "# md"),
        ("json", '{"type":"schema"}'),
    ):
        output = processor.process(
            _request(),
            content,
            media_kind=media_kind,
            content_manifest_id="manifest_1",
            content_manifest_hash="manifest_hash_1",
        )
        assert output.receipt.processing_summary["route"]["mineru"] is False
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["name", "value"])
    sheet.append(["a", "1"])
    stream = io.BytesIO()
    workbook.save(stream)
    output = processor.process(
        _request(),
        stream.getvalue(),
        media_kind="xlsx",
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
    )
    assert output.receipt.processing_summary["route"]["structured_loader"] is True
    assert output.chunks[0].locator["sheet"] == "Sheet"
    csv_output = processor.process(
        _request(),
        "name,value\na,1",
        media_kind="text/csv",
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
    )
    assert csv_output.chunks[0].locator == {"sheet": "CSV", "a1_range": "A2:B2"}


def test_a2_ocr_threshold_change_persists_new_boolean_score_and_version() -> None:
    facts = _pdf_facts(ocr_confidence_by_page={1: 0.7, 2: 0.7}, ocr_confidence=0.7)
    low = ContentProcessor(mineru=lambda content: facts, ocr_confidence_threshold=0.95).process(
        _request(),
        b"pdf",
        media_kind="application/pdf",
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
        page_count=2,
    )
    high = ContentProcessor(mineru=lambda content: facts, ocr_confidence_threshold=0.5).process(
        _request(),
        b"pdf",
        media_kind="application/pdf",
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
        page_count=2,
    )
    assert low.receipt.ocr_low_confidence is True
    assert high.receipt.ocr_low_confidence is False
    assert low.receipt.ocr_low_confidence_fact == {
        "confidence": 0.7,
        "page": 1,
        "region": [],
    }
    assert low.receipt.processing_summary["ocr"]["threshold_version"] == "ocr_threshold:0.95"
    assert high.receipt.processing_summary["ocr"]["threshold_version"] == "ocr_threshold:0.5"
    assert low.receipt.processing_summary["ocr"]["local_confidence"] == {"1": 0.7, "2": 0.7}


def test_a3_structure_classification_fixtures_are_mutually_exclusive() -> None:
    fixtures = {
        "tree": "# Chapter One\nbody\n\n## Section\nbody",
        "partial": "intro without heading\n\n## Skipped Level\nbody",
        "basic": "plain body without any structure signal",
    }
    assert classify_structure(fixtures["tree"], sample_pages=[1])["class"] == "tree"
    assert classify_structure(fixtures["partial"], sample_pages=[1])["class"] == "partial"
    assert classify_structure(fixtures["basic"], sample_pages=[1])["class"] == "basic"
    tree = ContentProcessor(
        mineru=lambda content: _pdf_facts(text=fixtures["tree"], structure={"class": "tree"})
    ).process(
        _request(),
        b"pdf",
        media_kind="application/pdf",
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
        page_count=1,
    )
    assert tree.receipt.processing_summary["tree"]["structure_class"] == "tree"
    partial = ContentProcessor(
        mineru=lambda content: _pdf_facts(text=fixtures["partial"], structure={"class": "partial"})
    ).process(
        _request(),
        b"pdf",
        media_kind="application/pdf",
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
        page_count=1,
    )
    assert partial.chunks[0].metadata["section_path"] == "此章未能分类"
    basic = ContentProcessor(
        mineru=lambda content: _pdf_facts(text=fixtures["basic"], structure={"class": "basic"})
    ).process(
        _request(),
        b"pdf",
        media_kind="application/pdf",
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
        page_count=1,
    )
    assert basic.receipt.processing_summary["tree"] == {
        "tree_indexed": False,
        "tree_reason": "no_structure",
        "structure_class": "basic",
    }
    assert all(chunk.locator == {} for chunk in basic.chunks)


def test_a3_pipeline_failure_returns_mineru_parse_failed_without_receipt() -> None:
    def failing_runner(*args: Any, **kwargs: Any) -> subprocess.CompletedProcess[bytes]:
        raise subprocess.CalledProcessError(1, ["mineru"])

    adapter = MinerUAdapter(runner=failing_runner)
    with pytest.raises(PlatformError) as error:
        adapter.parse(b"pdf", media_kind="application/pdf", page_count=2)
    assert error.value.code == "mineru_parse_failed"
    with pytest.raises(PlatformError) as processor_error:
        ContentProcessor(mineru=adapter).process(
            _request(),
            b"pdf",
            media_kind="application/pdf",
            content_manifest_id="manifest_1",
            content_manifest_hash="manifest_hash_1",
            page_count=2,
        )
    assert processor_error.value.code == "mineru_parse_failed"


def _capture_transport(payloads: list[dict[str, Any]], response: dict[str, Any]):
    def transport(url, payload, headers, options):
        payloads.append(
            {"url": url, "payload": json.loads(payload), "headers": headers, "options": options}
        )
        return response

    return transport


def test_a5_bailian_and_internvl_receive_context_and_ocr_tag_and_return_typed_result() -> None:
    payloads: list[dict[str, Any]] = []
    response = {
        "choices": [{"message": {"content": "A revenue chart"}}],
        "usage": {"input_tokens": 10, "output_tokens": 5},
    }
    usage_facts: list[dict[str, Any]] = []
    bailian = BailianImageDescriber(
        base_url="https://bailian.example",
        api_key="secret",
        model="qwen-vl-plus",
        usage_sink=usage_facts.append,
        transport=_capture_transport(payloads, response),
    )
    result = bailian(
        b"image-bytes",
        {
            "caption": "Revenue chart",
            "section_path": "2 / Results",
            "preceding_text": "Revenue rose.",
            "following_text": "The chart summarizes.",
            "ocr_text": "Q1: 10",
        },
    )
    prompt = payloads[0]["payload"]["messages"][0]["content"][0]["text"]
    assert "caption: Revenue chart" in prompt
    assert "section_path: 2 / Results" in prompt
    assert "ocr_tag: Q1: 10" in prompt
    assert payloads[0]["payload"]["messages"][0]["content"][1]["type"] == "image_url"
    assert result.text == "A revenue chart"
    assert result.indexable is True and result.degraded is False
    assert result.usage is not None and result.usage["latency_ms"] >= 0
    assert usage_facts[0]["kind"] == "provider_usage"
    internvl = InternVLImageDescriber(
        base_url="https://internvl.example",
        model="InternVL-3",
        revision="r1",
        usage_sink=usage_facts.append,
        transport=_capture_transport(payloads, response),
    )
    internvl_result = internvl(b"image-bytes", {"caption": "C", "ocr_text": "T"})
    assert internvl_result.provider == "internvl"
    assert usage_facts[-1]["kind"] == "local_usage"
    assert usage_facts[-1]["stage"] == "image_vlm_internvl"
    output = ContentProcessor(image_describer=bailian).process(
        _request(),
        b"image-bytes",
        media_kind="image/png",
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
        image_context={"caption": "Revenue chart"},
        image_ocr_text="Q1: 10",
    )
    assert "image-bytes" not in output.chunks[0].text
    assert output.chunks[0].text == "Revenue chart\nA revenue chart\nQ1: 10"


def test_a6_decorative_filter_none_provider_and_quota_usage_separation() -> None:
    assert is_decorative(50, 80) is True
    assert is_decorative(100, 100, page_area=1_000_000) is True
    assert is_decorative(500, 500, page_area=1_000_000) is False
    png = bytes.fromhex("89504e470d0a1a0a0000000d49484452000000" "3200000020080600000" "0")
    assert image_dimensions(png) == (50, 32)
    called: list[bytes] = []

    def describer(content: bytes, context: dict[str, Any]) -> ImageDescriptionResult:
        called.append(content)
        raise AssertionError("decorative image must not call VLM")

    output = ContentProcessor(image_describer=describer).process(
        _request(),
        png,
        media_kind="image/png",
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
    )
    assert called == []
    assert output.chunks == ()
    assert output.receipt.degradations[0]["reason"] == "decorative"
    assert output.receipt.processing_summary["filtered_image_count"] == 1
    none_result = NoneImageDescriber()(b"img", {"caption": "Cap", "ocr_text": "OCR text"})
    assert none_result.degraded is True
    assert none_result.indexable is True
    assert none_result.text == "Cap\nOCR text"
    empty = NoneImageDescriber()(b"img", {})
    assert empty.indexable is False
    assert empty.reason == "degraded_no_text"
    with pytest.raises(PlatformError) as error:
        NoneImageDescriber(environment="production")
    assert error.value.code == "image_provider_unavailable"


def _xlsx_bytes(workbook: Workbook) -> bytes:
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def test_a7_a8_multilevel_headers_merged_cells_and_total_rows() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "财务数据"
    sheet.merge_cells("A1:C1")
    sheet["A2"] = "营收"
    sheet["B2"] = "利润"
    sheet["C2"] = "季度"
    sheet.merge_cells("A2:A3")
    sheet.append([])
    sheet["A4"] = "营收"
    sheet["B4"] = "100"
    sheet["C4"] = "Q1"
    sheet["D1"] = "备注"
    sheet["D4"] = "ok"
    output = ContentProcessor().process(
        _request(),
        _xlsx_bytes(workbook),
        media_kind="xlsx",
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
    )
    headers = output.chunks[0].metadata["headers"]
    assert "财务数据_营收" in headers
    assert "财务数据_利润" in headers
    assert "财务数据_季度" in headers
    assert output.chunks[0].locator["sheet"] == "Sheet"


def test_a9_continuation_tables_merge_before_splitting_with_header_dedup() -> None:
    headers = ["区域", "销售额", "季度"]
    first = {
        "page": 1,
        "headers": headers,
        "rows": [headers, ["华东", "10", "Q1"], ["华南", "20", "Q1"]],
    }
    second = {
        "page": 2,
        "headers": headers,
        "rows": [headers, ["华北", "30", "Q1"]],
    }
    separate = {
        "page": 3,
        "headers": ["不同", "列"],
        "rows": [["不同", "列"], ["x", "y"]],
    }
    merged, facts = merge_continuation_tables([first, second, separate])
    assert len(merged) == 2
    assert merged[0]["rows"] == [["华东", "10", "Q1"], ["华南", "20", "Q1"], ["华北", "30", "Q1"]]
    assert merged[0]["page_start"] == 1 and merged[0]["page_end"] == 2
    assert facts == [{"kind": "continuation_header_deduplicated", "pages": [1, 2]}]


def test_a10_a11_markdown_tables_stay_whole_and_row_groups_follow_category() -> None:
    headers = ["区域", "销售额"]
    rows = [["华东", str(index)] for index in range(40)]
    output = ContentProcessor(
        mineru=lambda content: _pdf_facts(tables=[{"page": 1, "headers": headers, "rows": rows}])
    ).process(
        _request(),
        b"pdf",
        media_kind="application/pdf",
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
        page_count=1,
    )
    table_chunks = [chunk for chunk in output.chunks if chunk.media_kind == "table/markdown"]
    assert len(table_chunks) == 2
    for chunk in table_chunks:
        lines = chunk.text.splitlines()
        assert lines[0].startswith("| 区域 | 销售额 |")
        assert lines[1].startswith("| --- | --- |")
        assert all(line.startswith("|") for line in lines)
        assert chunk.metadata["headers"] == headers
        assert "region" not in chunk.metadata
    assert table_chunks[0].metadata["block"] == 1
    assert table_chunks[0].metadata["total_blocks"] == 2
    assert table_chunks[0].locator == {"page": 1}
    grouped = [["a", "1"], ["a", "2"], ["b", "3"], ["b", "4"]]
    assert _split_rows(grouped, 2) == [(0, 2), (2, 4)]
    wide = [["x"] * 25 for _ in range(45)]
    assert _split_rows(wide, 25)[0] == (0, 20)


def test_a12_a13_locators_and_snippets_follow_media_kind_contract() -> None:
    pdf = ContentProcessor(mineru=lambda content: _pdf_facts()).process(
        _request(),
        b"pdf",
        media_kind="application/pdf",
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
        page_count=2,
    )
    assert pdf.chunks[0].locator == {"page": 1, "span": "0:5"}
    assert pdf.chunks[0].snippet is not None
    text = ContentProcessor().process(
        _request(),
        "# Section\nbody",
        media_kind="text/markdown",
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
    )
    assert text.chunks[0].locator == {"section_path": "Section"}
    assert text.chunks[0].snippet is not None
    image = ContentProcessor(
        image_describer=lambda content, context: ImageDescriptionResult(
            text="desc", indexable=True, degraded=False, reason="ok", provider="bailian"
        )
    ).process(
        _request(),
        b"image-bytes",
        media_kind="image/png",
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
    )
    assert image.chunks[0].locator == {}
    assert image.chunks[0].snippet is None
    assert image.receipt.locator_snippet_integrity["locators_valid"] is True


def test_a14_table_parse_failure_fails_whole_attempt() -> None:
    with pytest.raises(PlatformError) as error:
        ContentProcessor().process(
            _request(),
            b"not-an-excel",
            media_kind="xlsx",
            content_manifest_id="manifest_1",
            content_manifest_hash="manifest_hash_1",
        )
    assert error.value.code == "table_parse_failed"


def test_unsupported_media_type_is_stable() -> None:
    with pytest.raises(PlatformError) as error:
        ContentProcessor().process(
            _request(),
            b"content",
            media_kind="application/x-unknown",
            content_manifest_id="manifest_1",
            content_manifest_hash="manifest_hash_1",
        )
    assert error.value.code == "unsupported_media_type"


def test_mineru_and_vlm_provider_settings_parse_from_environment(tmp_path) -> None:
    settings = load_platform_settings(
        {
            "RAG_PLATFORM_PROFILE": "development",
            "RAG_DATABASE_URL": "sqlite+pysqlite:///:memory:",
            "RAG_OBJECT_STORAGE_ENDPOINT": "http://localhost:9000",
            "RAG_OBJECT_STORAGE_BUCKET": "rag-dev",
            "RAG_PROVIDER_NAME": "fake",
            "RAG_INDEX_MINERU_PROVIDER": "local",
            "RAG_INDEX_MINERU_EXECUTABLE": "mineru-dev",
            "RAG_INDEX_IMAGE_VLM_PROVIDER": "bailian",
            "RAG_INDEX_IMAGE_VLM_BASE_URL": "https://bailian.example/v1",
            "RAG_INDEX_IMAGE_VLM_API_KEY": "secret",
            "RAG_INDEX_IMAGE_VLM_MODEL": "qwen-vl-max",
        }
    )
    assert settings.index.mineru_provider == "local"
    assert settings.index.mineru_executable == "mineru-dev"
    assert settings.index.image_vlm_provider == "bailian"
    assert settings.index.image_vlm_model == "qwen-vl-max"


def test_production_bailian_vlm_requires_base_url_and_api_key(tmp_path) -> None:
    base = {
        "RAG_PLATFORM_PROFILE": "production",
        "RAG_DATABASE_URL": "postgresql+psycopg://u:p@localhost/db",
        "RAG_OBJECT_STORAGE_ENDPOINT": "https://s3.example.com",
        "RAG_OBJECT_STORAGE_BUCKET": "rag",
        "RAG_PROVIDER_NAME": "dashscope",
        "RAG_PROVIDER_API_KEY": "secret",
        "RAG_AUTH_SECRET_KEY": "secret-key-long-enough",
        "RAG_AUTH_ALLOWED_ORIGINS": "https://app.example.com",
        "RAG_AUTH_ADMIN_ROSTER": "root",
        "RAG_BUSINESS_TIMEZONE": "UTC",
        "USER_DELETION_ARCHIVE_DIR": str(tmp_path / "archive"),
        "RAG_INDEX_IMAGE_VLM_PROVIDER": "bailian",
    }
    with pytest.raises(ValueError, match="base URL"):
        load_platform_settings(dict(base))
    with pytest.raises(ValueError, match="API key"):
        load_platform_settings({**base, "RAG_INDEX_IMAGE_VLM_BASE_URL": "https://b.example"})
