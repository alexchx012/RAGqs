from __future__ import annotations

import io
import json
import zipfile
from dataclasses import replace
from datetime import UTC, datetime, timedelta
from hashlib import sha256

import pytest
from openpyxl import Workbook
from sqlalchemy import create_engine, select, update

from app.documents.indexing import IndexProcessingReceipt, IndexStagingRequest
from app.documents.preview import ProcessingReceiptPreviewRenderer
from app.documents.public_graph import PublicGraphSourceService
from app.documents.schema import (
    document_versions_table,
    documents_instance_counters_table,
    documents_metadata,
    documents_table,
    index_changes_table,
    index_revisions_table,
    publications_table,
)
from app.documents.service import DocumentUpload
from app.indexing import (
    ContentProcessor,
    DocumentVisibilityFact,
    GenerationManager,
    GraphComponentCoordinator,
    IndexChunk,
    IndexGenerationComponentInput,
    IndexingService,
    InMemoryIndexWriter,
    InMemorySparseIndexProvider,
    NarrowingScope,
    RetrievalProfile,
    RetrievalReleaseService,
    RetrievalScope,
    RetrievalService,
    SqlAlchemyGenerationManager,
    SqlAlchemyIndexingRepository,
    indexing_metadata,
    intersect_scopes,
)
from app.indexing.models import RetrievalHit
from app.indexing.processing import _xlsx_merged_ranges
from app.indexing.retrieval import CitationService
from app.indexing.schema import index_chunks_table, index_generation_heads_table
from app.platform.errors import PlatformError
from app.platform.storage import MemoryObjectStore, ObjectMetadata


class _Outbox:
    def publish_public_graph_source_change(self, **event: object) -> str:
        del event
        return "event_1"


def _graph_input(snapshot, *, operation_id: str, target_generation_id: str = "generation_graph"):
    return IndexGenerationComponentInput(
        component_kind="public_graph",
        target_generation_id=target_generation_id,
        target_generation_fence="generation_fence_1",
        source_snapshot=snapshot,
        source_manifest_hash=snapshot.source_manifest_hash,
        source_head_fence=1,
        operation_id=operation_id,
    )


def _request(
    *, generation: str = "generation_initial", attempt_id: str = "attempt_1"
) -> IndexStagingRequest:
    return IndexStagingRequest(
        job_id="job_1",
        attempt_id=attempt_id,
        fencing_token=1,
        publication_id="publication_1",
        document_id="document_1",
        document_version_id="version_1",
        space_id="space_1",
        operation="initial",
        base_active_version_id=None,
        expected_generation_id=generation,
        index_revision_at_start=0,
        object_manifest_ref="manifest_1",
        processing_config_snapshot={},
        authorization_fence={"actor_id": "user_1"},
        input_manifest_hash="manifest_hash_1",
        processing_profile_version="profile_1",
    )


def _chunk(
    chunk_id: str,
    *,
    generation: str = "generation_initial",
    document_id: str = "document_1",
) -> IndexChunk:
    return IndexChunk(
        chunk_id=chunk_id,
        generation_id=generation,
        publication_id="publication_1",
        document_id=document_id,
        document_version_id="version_1",
        space_id="space_1",
        text=f"text {chunk_id}",
        embedding_text=f"text {chunk_id}",
        locator={"section_path": chunk_id},
        snippet=f"snippet {chunk_id}",
        media_kind="text/plain",
        manifest_hash="manifest_hash_1",
    )


def _release_suite() -> dict[str, object]:
    return {
        "acl_assertions": {"space_isolation": "passed"},
        "hardware_profile": {"accelerator": "test"},
        "thresholds": {
            "p50_ms": 10,
            "p95_ms": 20,
            "p99_ms": 30,
            "error_rate": 0.01,
            "vram_mb": 100,
        },
        "samples": {
            name: [{"sample_id": f"{name}-1", "input": name, "expected": "pass"}]
            for name in (
                "phrase_query",
                "proper_noun_query",
                "quoted_exact_query",
                "real_question",
                "acl_filter",
                "sparse_exact_hit",
                "refusal",
            )
        },
        "quality_thresholds": {"hit_at_k": 0.8, "mrr": 0.8, "ndcg": 0.8, "refusal": 0.9},
    }


def _release_metrics() -> dict[str, float]:
    return {
        "p50_ms": 1,
        "p95_ms": 2,
        "p99_ms": 3,
        "error_rate": 0,
        "vram_mb": 1,
        "hit_at_k": 0.9,
        "mrr": 0.9,
        "ndcg": 0.9,
        "refusal": 1.0,
    }


def _insert_active_publication(
    connection,
    *,
    content: bytes,
    document_id: str = "document_1",
    document_version_id: str = "version_1",
    publication_id: str = "publication_1",
    space_id: str = "space_1",
    object_key: str = "documents/document_1/version_1/original",
) -> str:
    now = datetime(2026, 1, 1, tzinfo=UTC)
    content_hash = sha256(content).hexdigest()
    connection.execute(
        documents_table.insert().values(
            id=document_id,
            space_id=space_id,
            lifecycle_status="active",
            active_version_id=document_version_id,
            pending_version_id=None,
            active_operation_job_id=None,
            deletion_id=None,
            version=1,
            name="document.md",
            normalized_name="document.md",
            media_kind="text/markdown",
            uploaded_at_utc=now,
            created_by_user_id="user_1",
            created_at_utc=now,
            updated_at_utc=now,
        )
    )
    connection.execute(
        document_versions_table.insert().values(
            id=document_version_id,
            document_id=document_id,
            version_number=1,
            status="active",
            content_hash_sha256=content_hash,
            object_manifest_json={"object_key": object_key, "size_bytes": len(content)},
            original_object_key=object_key,
            file_name="document.md",
            media_kind="text/markdown",
            size_bytes=len(content),
            created_by_user_id="user_1",
            activated_at_utc=now,
            terminal_at_utc=None,
            superseded_at_utc=None,
            purge_after_at_utc=None,
            purged_at_utc=None,
            restored_from_version_id=None,
            created_at_utc=now,
            updated_at_utc=now,
        )
    )
    connection.execute(
        publications_table.insert().values(
            id=publication_id,
            document_id=document_id,
            document_version_id=document_version_id,
            job_id=f"job_{document_id}",
            attempt_id=f"attempt_{document_id}",
            generation_id="generation_initial",
            status="active",
            resource_manifest_json={
                "content_manifest_id": f"manifest_{publication_id}",
                "content_manifest_hash": content_hash,
                "processing_profile_version": "profile_1",
                "processing_config_snapshot": {},
            },
            created_at_utc=now,
            activated_at_utc=now,
            superseded_at_utc=None,
            discarded_at_utc=None,
        )
    )
    return content_hash


def test_scope_intersection_cannot_expand_server_scope() -> None:
    allowed = RetrievalScope(
        frozenset({"space_1", "space_2"}),
        documents_by_space={"space_1": frozenset({"doc_1"}), "space_2": frozenset({"doc_2"})},
    )
    effective = intersect_scopes(
        allowed,
        NarrowingScope(frozenset({"space_1", "space_3"}), frozenset({"doc_2", "doc_3"})),
    )
    assert effective.space_ids == frozenset({"space_1"})
    assert effective.is_empty
    assert not effective.allows(space_id="space_2", document_id="doc_2")


def test_retrieval_replenishes_cursor_after_visibility_filter() -> None:
    provider = InMemorySparseIndexProvider()
    provider.stage_chunks(
        "attempt_1",
        "publication_1",
        "document_1",
        "version_1",
        [_chunk("chunk_1"), _chunk("chunk_2")],
    )
    provider.publish_staged("attempt_1", "publication_1")
    manager = GenerationManager()

    def facts(candidate: IndexChunk, principal: object) -> DocumentVisibilityFact:
        del principal
        return DocumentVisibilityFact(
            candidate.document_id,
            candidate.space_id,
            "superseded" if candidate.chunk_id == "chunk_1" else "active",
            candidate.document_version_id,
            candidate.publication_id,
            "active",
            candidate.manifest_hash,
            True,
        )

    service = RetrievalService(
        manager,
        [provider],
        identity_access=lambda principal: RetrievalScope(frozenset({"space_1"})),
        visibility_facts=facts,
    )
    result = service.search(
        "text",
        principal="user_1",
        profile=RetrievalProfile(top_k=1, candidate_limit=1),
    )
    assert [hit.chunk.chunk_id for hit in result.hits] == ["chunk_2"]


def test_retrieval_requires_authoritative_visibility_facts() -> None:
    provider = InMemorySparseIndexProvider()
    provider.stage_chunks(
        "attempt_1",
        "publication_1",
        "document_1",
        "version_1",
        [_chunk("chunk_1")],
    )
    provider.publish_staged("attempt_1", "publication_1")
    service = RetrievalService(
        GenerationManager(),
        [provider],
        identity_access=lambda principal: RetrievalScope(frozenset({"space_1"})),
    )

    result = service.search("text", principal="user_1")

    assert result.hits == ()


def test_retrieval_resolver_ignores_caller_profile_overrides() -> None:
    provider = InMemorySparseIndexProvider()
    manager = GenerationManager()
    requested = RetrievalProfile(
        profile_id="released",
        version="7",
        top_k=1,
        candidate_limit=1,
        effort="deep",
        route_tree=True,
    )
    released = RetrievalProfile(
        profile_id="released",
        version="7",
        top_k=4,
        candidate_limit=12,
        effort="quick",
        release_id="release_7",
    )
    seen: list[RetrievalProfile] = []
    service = RetrievalService(
        manager,
        [provider],
        profile_resolver=lambda profile, generation_id: seen.append(profile) or released,
    )

    result = service.search("text", profile=requested)

    assert seen == [RetrievalProfile(profile_id="released", version="7")]
    assert result.profile == released


def test_retrieval_keeps_hybrid_before_rerank_and_routes_tree_afterwards() -> None:
    dense = InMemoryIndexWriter(provider_name="dense")
    sparse = InMemorySparseIndexProvider(provider_name="sparse")
    for provider, chunk_id in ((dense, "dense_1"), (sparse, "sparse_1")):
        document_id = f"document_{provider.provider_name}"
        provider.stage_chunks(
            provider.provider_name,
            "publication_1",
            document_id,
            "version_1",
            [_chunk(chunk_id, document_id=document_id)],
        )
        provider.publish_staged(provider.provider_name, "publication_1")

    events: list[tuple[str, object]] = []

    class Reranker:
        def rerank(self, query, hits, profile):
            del query, profile
            events.append(("rerank", tuple(hit.chunk.chunk_id for hit in hits)))
            return tuple(reversed(hits)), None

    def tree_router(query, candidates, *, max_documents, rag_call_limit):
        del query
        events.append(
            (
                "tree",
                (tuple(hit.chunk.chunk_id for hit in candidates), max_documents, rag_call_limit),
            )
        )

    service = RetrievalService(
        GenerationManager(),
        [dense, sparse],
        identity_access=lambda principal: RetrievalScope(frozenset({"space_1"})),
        visibility_facts=lambda candidate, principal: DocumentVisibilityFact(
            candidate.document_id,
            candidate.space_id,
            "active",
            candidate.document_version_id,
            candidate.publication_id,
            "active",
            candidate.manifest_hash,
            True,
        ),
        reranker=Reranker(),
        tree_router=tree_router,
    )

    result = service.search(
        "text",
        principal="user_1",
        profile=RetrievalProfile(
            top_k=2,
            candidate_limit=2,
            effort="think",
            route_tree=True,
        ),
    )

    assert events == [
        ("rerank", ("dense_1", "sparse_1")),
        # Equal scores fall back to the chunk_id tie-break, so the reranker's
        # reversed output still routes in a deterministic order (A6).
        ("tree", (("dense_1", "sparse_1"), 7, 4)),
    ]
    assert [hit.chunk.chunk_id for hit in result.hits] == ["dense_1", "sparse_1"]


def test_cleanup_resource_removes_dense_and_sparse_document_version_resources() -> None:
    dense = InMemoryIndexWriter(provider_name="dense")
    sparse = InMemorySparseIndexProvider(provider_name="sparse")
    chunk = replace(
        _chunk("cleanup_chunk", document_id="document_cleanup"),
        publication_id="publication_cleanup",
        document_version_id="version_cleanup",
    )
    for provider in (dense, sparse):
        provider.stage_chunks(
            "attempt_cleanup",
            "publication_cleanup",
            "document_cleanup",
            "version_cleanup",
            [chunk],
        )
        provider.publish_staged("attempt_cleanup", "publication_cleanup")
    service = IndexingService(dense_writer=dense, sparse_provider=sparse)

    service.cleanup_resource(
        {
            "backend_kind": "index_chunk",
            "resource_id": "cleanup_chunk",
            "document_id": "document_cleanup",
            "document_version_id": "version_cleanup",
        }
    )

    assert dense.visible_chunks() == ()
    assert sparse.visible_chunks() == ()


@pytest.mark.parametrize(
    ("effort", "expected_calls", "expected_documents", "tree_expected"),
    (("quick", 1, 5, False), ("think", 4, 7, True), ("deep", 10, 9, True)),
)
def test_retrieval_effort_sets_routing_budgets(
    effort: str,
    expected_calls: int,
    expected_documents: int,
    tree_expected: bool,
) -> None:
    tree_calls: list[tuple[int, int]] = []
    graph_calls: list[int] = []
    service = RetrievalService(
        GenerationManager(),
        [InMemorySparseIndexProvider()],
        identity_access=lambda principal: RetrievalScope(frozenset({"space_1"})),
        graph_reader=type(
            "Reader",
            (),
            {
                "acquire_current_reader_lease": lambda self, *, generation_id: object(),
                "release_reader_lease": lambda self, lease: None,
            },
        )(),
        tree_router=lambda query, candidates, *, max_documents, rag_call_limit: tree_calls.append(
            (max_documents, rag_call_limit)
        ),
        graph_router=lambda query, candidates, *, rag_call_limit, reader_lease: graph_calls.append(
            rag_call_limit
        ),
    )

    service.search(
        "text",
        profile=RetrievalProfile(effort=effort, route_tree=True, route_graph=True),
    )

    assert tree_calls == ([(expected_documents, expected_calls)] if tree_expected else [])
    assert graph_calls == [expected_calls]


def test_graph_route_degradation_keeps_hybrid_retrieval_result() -> None:
    provider = InMemorySparseIndexProvider()
    provider.stage_chunks(
        "attempt_1",
        "publication_1",
        "document_1",
        "version_1",
        [_chunk("chunk_1")],
    )
    provider.publish_staged("attempt_1", "publication_1")
    service = RetrievalService(
        GenerationManager(),
        [provider],
        identity_access=lambda principal: RetrievalScope(frozenset({"space_1"})),
        visibility_facts=lambda candidate, principal: DocumentVisibilityFact(
            candidate.document_id,
            candidate.space_id,
            "active",
            candidate.document_version_id,
            candidate.publication_id,
            "active",
            candidate.manifest_hash,
            True,
        ),
        graph_reader=type(
            "Reader",
            (),
            {
                "acquire_current_reader_lease": lambda self, *, generation_id: object(),
                "release_reader_lease": lambda self, lease: None,
            },
        )(),
        graph_router=lambda query, candidates, *, rag_call_limit, reader_lease: (
            _ for _ in ()
        ).throw(PlatformError("graph_stale", "graph is stale", {}, 409)),
    )

    result = service.search(
        "text",
        principal="user_1",
        profile=RetrievalProfile(route_graph=True),
    )

    assert [hit.chunk.chunk_id for hit in result.hits] == ["chunk_1"]
    assert result.degradations[-1] == {"code": "graph_degraded", "reason": "graph_stale"}


def test_missing_graph_router_is_an_explicit_degradation() -> None:
    provider = InMemorySparseIndexProvider()
    provider.stage_chunks(
        "attempt_1",
        "publication_1",
        "document_1",
        "version_1",
        [_chunk("chunk_1")],
    )
    provider.publish_staged("attempt_1", "publication_1")
    service = RetrievalService(
        GenerationManager(),
        [provider],
        identity_access=lambda principal: RetrievalScope(frozenset({"space_1"})),
        visibility_facts=lambda candidate, principal: DocumentVisibilityFact(
            candidate.document_id,
            candidate.space_id,
            "active",
            candidate.document_version_id,
            candidate.publication_id,
            "active",
            candidate.manifest_hash,
            True,
        ),
    )

    result = service.search(
        "text",
        principal="user_1",
        profile=RetrievalProfile(route_graph=True),
    )

    assert [hit.chunk.chunk_id for hit in result.hits] == ["chunk_1"]
    assert result.degradations[-1] == {
        "code": "graph_degraded",
        "reason": "graph_unavailable",
    }


def test_citation_withholds_locator_when_server_scope_excludes_document() -> None:
    candidate = _chunk("chunk_1")

    def facts(value: IndexChunk, principal: object) -> DocumentVisibilityFact:
        del principal
        return DocumentVisibilityFact(
            value.document_id,
            value.space_id,
            "active",
            value.document_version_id,
            value.publication_id,
            "active",
            value.manifest_hash,
            True,
        )

    citation = CitationService(
        facts,
        GenerationManager(),
        identity_access=lambda principal: RetrievalScope(frozenset()),
    )

    assert citation.resolve(RetrievalHit(candidate, 1.0, "dense"), principal="other") == {
        "state": "unavailable"
    }


def test_citation_uses_and_releases_a_generation_reference_lease() -> None:
    candidate = _chunk("chunk_1")

    class GenerationLeases:
        active_generation_id = "generation_initial"

        def __init__(self) -> None:
            self.acquired = 0
            self.released: list[str] = []

        def acquire_reference_lease(self):
            self.acquired += 1
            return type(
                "Lease",
                (),
                {"lease_id": "lease_1", "generation_id": "generation_initial"},
            )()

        def release_reference_lease(self, lease_id: str) -> None:
            self.released.append(lease_id)

    manager = GenerationLeases()
    citation = CitationService(
        lambda value, principal: DocumentVisibilityFact(
            value.document_id,
            value.space_id,
            "active",
            value.document_version_id,
            value.publication_id,
            "active",
            value.manifest_hash,
            True,
        ),
        manager,
        identity_access=lambda principal: RetrievalScope(frozenset({"space_1"})),
    )

    result = citation.resolve(RetrievalHit(candidate, 1.0, "dense"), principal="user_1")

    assert result["state"] == "available"
    assert manager.acquired == 1
    assert manager.released == ["lease_1"]


def test_generation_requires_contiguous_revision_before_release_and_supports_rollback() -> None:
    clock_value = [datetime(2026, 1, 1, tzinfo=UTC)]
    manager = GenerationManager(now=lambda: clock_value[0])
    staging = manager.create_staging([], base_revision=0)
    with pytest.raises(PlatformError) as error:
        manager.apply_change(staging.generation_id, 2, {"change_type": "publish"})
    assert error.value.code == "revision_gap"
    manager.apply_change(staging.generation_id, 1, {"change_type": "publish"})
    manager.set_component_state(staging.generation_id, "dense", "ready")
    active = manager.release(staging.generation_id, current_revision=1)
    assert manager.active_generation_id == active.generation_id
    manager.set_current_revision(2)
    retired = next(item for item in manager.list_generations() if item.status == "retired")
    manager.apply_change(retired.generation_id, 1, {"change_type": "publish"})
    manager.apply_change(retired.generation_id, 2, {"change_type": "publish"})
    clock_value[0] += timedelta(hours=1)
    restored = manager.rollback(
        retired.generation_id,
        current_revision=2,
        source_receipt={
            "state": "held",
            "candidate_generation_id": retired.generation_id,
            "applied_revision": 2,
        },
    )
    assert restored.generation_id == retired.generation_id


def test_sql_generation_head_and_operations_survive_repository_restart() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    documents_metadata.create_all(engine)
    indexing_metadata.create_all(engine)
    first = SqlAlchemyIndexingRepository(engine)
    staging = first.create_staging([], base_revision=0, generation_id="generation_next")
    with engine.begin() as connection:
        connection.execute(
            documents_instance_counters_table.insert().values(
                counter_name="index_revision", value=0
            )
        )
        connection.execute(
            update(documents_instance_counters_table)
            .where(documents_instance_counters_table.c.counter_name == "index_revision")
            .values(value=1)
        )
        connection.execute(
            index_revisions_table.insert().values(
                id="revision_1",
                document_id="document_1",
                revision=1,
                generation_id="generation_initial",
                created_at_utc=datetime(2026, 1, 1, tzinfo=UTC),
            )
        )
        connection.execute(
            index_changes_table.insert().values(
                id="change_1",
                document_id="document_1",
                document_version_id=None,
                publication_id=None,
                revision_id="revision_1",
                change_type="delete",
                space_id="space_1",
                created_at_utc=datetime(2026, 1, 1, tzinfo=UTC),
            )
        )
    released = first.release(staging.generation_id)
    assert released.generation_id == "generation_next"
    assert released.applied_revision == 1

    restarted = SqlAlchemyIndexingRepository(engine)
    assert restarted.active_generation_id() == "generation_next"
    assert restarted.get_generation("generation_initial").status == "retired"
    blocked = restarted.request_index_generation_gc(
        "generation_initial", reconciliation_run_id="reconcile_1", operation_id="gc_1"
    )
    assert blocked.state == "blocked"
    assert "rollback_candidate" in blocked.blocking_reasons


def test_sql_generation_catches_up_documents_changes_without_gaps() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    documents_metadata.create_all(engine)
    indexing_metadata.create_all(engine)
    repository = SqlAlchemyIndexingRepository(engine)
    staging = repository.create_staging([], base_revision=0, generation_id="generation_next")
    with engine.begin() as connection:
        _insert_active_publication(connection, content=b"document source")
        connection.execute(
            documents_instance_counters_table.insert().values(
                counter_name="index_revision", value=0
            )
        )
        connection.execute(
            update(documents_instance_counters_table)
            .where(documents_instance_counters_table.c.counter_name == "index_revision")
            .values(value=1)
        )
        connection.execute(
            index_revisions_table.insert().values(
                id="revision_1",
                document_id="document_1",
                revision=1,
                generation_id="generation_initial",
                created_at_utc=datetime(2026, 1, 1, tzinfo=UTC),
            )
        )
        connection.execute(
            index_changes_table.insert().values(
                id="change_1",
                document_id="document_1",
                document_version_id="version_1",
                publication_id="publication_1",
                revision_id="revision_1",
                change_type="publish",
                space_id="space_1",
                created_at_utc=datetime(2026, 1, 1, tzinfo=UTC),
            )
        )
    caught_up = repository.catch_up_from_documents(staging.generation_id)
    assert caught_up.applied_revision == 1


def test_generation_catch_up_rejects_a_publish_change_without_a_documents_source() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    documents_metadata.create_all(engine)
    indexing_metadata.create_all(engine)
    repository = SqlAlchemyIndexingRepository(engine)
    staging = repository.create_staging([], base_revision=0, generation_id="generation_next")
    with engine.begin() as connection:
        connection.execute(
            documents_instance_counters_table.insert().values(
                counter_name="index_revision", value=0
            )
        )
        connection.execute(
            update(documents_instance_counters_table)
            .where(documents_instance_counters_table.c.counter_name == "index_revision")
            .values(value=1)
        )
        connection.execute(
            index_revisions_table.insert().values(
                id="revision_1",
                document_id="document_1",
                revision=1,
                generation_id="generation_initial",
                created_at_utc=datetime(2026, 1, 1, tzinfo=UTC),
            )
        )
        connection.execute(
            index_changes_table.insert().values(
                id="change_1",
                document_id="document_1",
                document_version_id="version_1",
                publication_id="publication_1",
                revision_id="revision_1",
                change_type="publish",
                space_id="space_1",
                created_at_utc=datetime(2026, 1, 1, tzinfo=UTC),
            )
        )

    with pytest.raises(PlatformError) as error:
        repository.catch_up_from_documents(staging.generation_id)

    assert error.value.code == "generation_source_missing"


def test_staging_generation_builds_its_frozen_base_snapshot_before_release() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    documents_metadata.create_all(engine)
    indexing_metadata.create_all(engine)
    object_store = MemoryObjectStore()
    content = b"# Snapshot heading\n\nrebuildable content"
    object_key = "documents/document_1/version_1/original"
    object_store.put(
        object_key,
        content,
        ObjectMetadata(content_type="text/markdown", size_bytes=len(content)),
    )
    with engine.begin() as connection:
        _insert_active_publication(connection, content=content, object_key=object_key)

    repository = SqlAlchemyIndexingRepository(engine)
    manager = SqlAlchemyGenerationManager(repository)
    service = IndexingService(generation_manager=manager, object_store=object_store)

    staging = manager.create_staging((), generation_id="generation_snapshot")

    with engine.connect() as connection:
        chunks = (
            connection.execute(
                select(index_chunks_table).where(
                    index_chunks_table.c.generation_id == staging.generation_id
                )
            )
            .mappings()
            .all()
        )
    assert [(row["publication_id"], row["manifest_hash"]) for row in chunks] == [
        ("publication_1", sha256(content).hexdigest())
    ]
    assert repository.release(staging.generation_id).generation_id == staging.generation_id
    assert [chunk.generation_id for chunk in service.dense_writer.visible_chunks()] == [
        staging.generation_id
    ]
    assert [chunk.generation_id for chunk in service.sparse_provider.visible_chunks()] == [
        staging.generation_id
    ]


def test_generation_catch_up_rebuilds_a_published_revision_from_documents_content() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    documents_metadata.create_all(engine)
    indexing_metadata.create_all(engine)
    object_store = MemoryObjectStore()
    content = b"# Changed heading\n\nchanged publication content"
    object_key = "documents/document_1/version_1/original"
    object_store.put(
        object_key,
        content,
        ObjectMetadata(content_type="text/markdown", size_bytes=len(content)),
    )
    with engine.begin() as connection:
        _insert_active_publication(connection, content=content, object_key=object_key)
        connection.execute(
            documents_instance_counters_table.insert().values(
                counter_name="index_revision", value=0
            )
        )
        connection.execute(
            update(documents_instance_counters_table)
            .where(documents_instance_counters_table.c.counter_name == "index_revision")
            .values(value=1)
        )
        connection.execute(
            index_revisions_table.insert().values(
                id="revision_1",
                document_id="document_1",
                revision=1,
                generation_id="generation_initial",
                created_at_utc=datetime(2026, 1, 1, tzinfo=UTC),
            )
        )
        connection.execute(
            index_changes_table.insert().values(
                id="change_1",
                document_id="document_1",
                document_version_id="version_1",
                publication_id="publication_1",
                revision_id="revision_1",
                change_type="publish",
                space_id="space_1",
                created_at_utc=datetime(2026, 1, 1, tzinfo=UTC),
            )
        )

    repository = SqlAlchemyIndexingRepository(engine)
    manager = SqlAlchemyGenerationManager(repository)
    dense = _CapturingStageWriter(provider_name="dense")
    _service = IndexingService(
        dense_writer=dense,
        sparse_provider=InMemorySparseIndexProvider(),
        generation_manager=manager,
        object_store=object_store,
        now=lambda: datetime(2026, 1, 1, 12, tzinfo=UTC),
    )
    staging = manager.create_staging([], base_revision=0, generation_id="generation_change")

    caught_up = manager.catch_up_from_documents(staging.generation_id)

    assert caught_up.applied_revision == 1
    with engine.connect() as connection:
        rows = connection.execute(
            select(index_chunks_table.c.publication_id, index_chunks_table.c.manifest_hash).where(
                index_chunks_table.c.generation_id == staging.generation_id
            )
        ).all()
    assert rows == [("publication_1", sha256(content).hexdigest())]
    assert dense.stage_kwargs is not None
    usage_context = dense.stage_kwargs["usage_context"]
    assert usage_context.execution_kind == "index_maintenance"
    assert usage_context.execution_id == "generation-build:generation_change:publication_1"
    assert usage_context.attempt_id == "generation-build:generation_change:publication_1"
    assert usage_context.generation_id == "generation_change"
    assert usage_context.publication_id == "publication_1"
    assert usage_context.replay_generation == 0
    assert usage_context.ownership.actor_user_id == "system:indexing"
    assert usage_context.ownership.actor_role_snapshot == "ops"
    assert usage_context.ownership.quota_subject_user_id is None
    assert usage_context.ownership.cost_center_key == "system:indexing"
    assert usage_context.ownership.source_space_ids == ("space_1",)
    assert usage_context.deadline_utc == datetime(2026, 1, 1, 12, 15, tzinfo=UTC)


def test_configuration_change_creates_a_new_staging_generation_with_immutable_manifest() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    documents_metadata.create_all(engine)
    indexing_metadata.create_all(engine)
    first = SqlAlchemyIndexingRepository(
        engine,
        generation_configuration={
            "provider": "meilisearch",
            "engine": "meilisearch",
            "analyzer": "jieba",
            "pretokenizer_version": "v1",
            "schema_version": "index-chunks-v1",
        },
    )

    configured = first.ensure_configuration_staging()

    assert configured.status == "staging"
    assert configured.manifest["indexing_configuration"]["provider"] == "meilisearch"
    assert configured.manifest["indexing_configuration"]["schema_hash"]
    assert configured.manifest["indexing_configuration"]["config_hash"]
    first.release(configured.generation_id)

    changed = SqlAlchemyIndexingRepository(
        engine,
        generation_configuration={
            "provider": "opensearch",
            "engine": "opensearch",
            "analyzer": "standard",
            "pretokenizer_version": "v1",
            "schema_version": "index-chunks-v1",
        },
    )

    next_generation = changed.ensure_configuration_staging()

    assert next_generation.generation_id != configured.generation_id
    assert next_generation.status == "staging"
    assert next_generation.manifest["indexing_configuration"]["provider"] == "opensearch"
    assert next_generation.manifest["provider"] == "opensearch"
    assert next_generation.manifest["dimension"] is None
    assert next_generation.manifest["metric"] == "cosine"
    assert next_generation.manifest["model_revision"] == "configured"
    assert next_generation.manifest["sparse_schema_hash"]
    assert next_generation.manifest["implementation_config_hash"]
    assert next_generation.manifest["last_applied_index_change_id"] is None
    assert next_generation.manifest["published_at"] is None
    assert next_generation.manifest["rollback_candidate_until"] is None
    assert next_generation.manifest["gc_state"] == "staging"
    assert next_generation.manifest["components"]["sparse"]["tokenizer_revision"]


def test_generation_catch_up_removes_chunks_for_a_documents_delete_change() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    documents_metadata.create_all(engine)
    indexing_metadata.create_all(engine)
    repository = SqlAlchemyIndexingRepository(engine)
    staging = repository.create_staging([], base_revision=0, generation_id="generation_next")
    with engine.begin() as connection:
        repository.record_published_chunks(
            _request(generation=staging.generation_id),
            (_chunk("chunk_1", generation=staging.generation_id),),
            connection=connection,
        )
        connection.execute(
            documents_instance_counters_table.insert().values(
                counter_name="index_revision", value=0
            )
        )
        connection.execute(
            update(documents_instance_counters_table)
            .where(documents_instance_counters_table.c.counter_name == "index_revision")
            .values(value=1)
        )
        connection.execute(
            index_revisions_table.insert().values(
                id="revision_1",
                document_id="document_1",
                revision=1,
                generation_id="generation_initial",
                created_at_utc=datetime(2026, 1, 1, tzinfo=UTC),
            )
        )
        connection.execute(
            index_changes_table.insert().values(
                id="change_1",
                document_id="document_1",
                document_version_id=None,
                publication_id=None,
                revision_id="revision_1",
                change_type="delete",
                space_id="space_1",
                created_at_utc=datetime(2026, 1, 1, tzinfo=UTC),
            )
        )

    repository.catch_up_from_documents(staging.generation_id)

    with engine.connect() as connection:
        remaining = connection.execute(
            index_chunks_table.select().where(
                index_chunks_table.c.generation_id == staging.generation_id
            )
        ).all()
    assert remaining == []


def test_generation_catch_up_discards_deleted_publication_from_generation_providers() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    documents_metadata.create_all(engine)
    indexing_metadata.create_all(engine)
    object_store = MemoryObjectStore()
    content = b"# Deletion heading\n\ncontent to remove"
    object_key = "documents/document_1/version_1/original"
    object_store.put(
        object_key,
        content,
        ObjectMetadata(content_type="text/markdown", size_bytes=len(content)),
    )
    with engine.begin() as connection:
        _insert_active_publication(connection, content=content, object_key=object_key)

    repository = SqlAlchemyIndexingRepository(engine)
    manager = SqlAlchemyGenerationManager(repository)
    service = IndexingService(generation_manager=manager, object_store=object_store)
    active_request = _request(generation="generation_initial", attempt_id="active_attempt")
    active_chunk = _chunk("active_chunk", generation="generation_initial")
    for provider in (service.dense_writer, service.sparse_provider):
        provider.stage_chunks(
            active_request.attempt_id,
            active_request.publication_id,
            active_request.document_id,
            active_request.document_version_id,
            (active_chunk,),
            expected_generation_id=active_request.expected_generation_id,
        )
        provider.publish_staged(
            active_request.attempt_id,
            active_request.publication_id,
            expected_generation_id=active_request.expected_generation_id,
        )
    staging = manager.create_staging((), generation_id="generation_delete")
    with engine.begin() as connection:
        connection.execute(
            documents_table.update()
            .where(documents_table.c.id == "document_1")
            .values(lifecycle_status="deleted")
        )
        connection.execute(
            documents_instance_counters_table.insert().values(
                counter_name="index_revision", value=0
            )
        )
        connection.execute(
            update(documents_instance_counters_table)
            .where(documents_instance_counters_table.c.counter_name == "index_revision")
            .values(value=1)
        )
        connection.execute(
            index_revisions_table.insert().values(
                id="revision_1",
                document_id="document_1",
                revision=1,
                generation_id="generation_initial",
                created_at_utc=datetime(2026, 1, 1, tzinfo=UTC),
            )
        )
        connection.execute(
            index_changes_table.insert().values(
                id="change_1",
                document_id="document_1",
                document_version_id=None,
                publication_id=None,
                revision_id="revision_1",
                change_type="delete",
                space_id="space_1",
                created_at_utc=datetime(2026, 1, 1, tzinfo=UTC),
            )
        )

    manager.catch_up_from_documents(staging.generation_id)

    for provider in (service.dense_writer, service.sparse_provider):
        assert [chunk.generation_id for chunk in provider.visible_chunks()] == [
            "generation_initial"
        ]


def test_sql_generation_keeps_logical_chunk_ids_per_generation() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    documents_metadata.create_all(engine)
    indexing_metadata.create_all(engine)
    repository = SqlAlchemyIndexingRepository(engine)

    with engine.begin() as connection:
        repository.record_published_chunks(_request(), (_chunk("chunk_1"),), connection=connection)
        repository.record_published_chunks(
            _request(generation="generation_next"),
            (_chunk("chunk_1", generation="generation_next"),),
            connection=connection,
        )


def test_sql_generation_rolls_back_within_its_persisted_window() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    documents_metadata.create_all(engine)
    indexing_metadata.create_all(engine)
    repository = SqlAlchemyIndexingRepository(engine)
    staging = repository.create_staging([], base_revision=0, generation_id="generation_next")
    repository.release(staging.generation_id)

    restored = repository.rollback(
        "generation_initial",
        source_receipt={
            "state": "held",
            "candidate_generation_id": "generation_initial",
            "applied_revision": 0,
        },
    )

    assert restored.generation_id == "generation_initial"


def test_graph_component_replays_stage_and_release_after_restart() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    documents_metadata.create_all(engine)
    indexing_metadata.create_all(engine)
    source = PublicGraphSourceService(
        engine,
        trusted_consumers={"indexing": {"indexer_1"}},
        outbox_port=_Outbox(),
    )
    snapshot = source.record_source_change(
        space_id="public",
        document_id="document_1",
        change_type="publish",
        publications=[],
    )
    first = GraphComponentCoordinator(
        SqlAlchemyGenerationManager(SqlAlchemyIndexingRepository(engine)),
        source,
        consumer_id="indexer_1",
    )
    grant = first.reserve_graph_component_stage(
        graph_build_id="graph_1",
        expected_source_revision=snapshot.source_revision,
        expected_active_generation_id="generation_initial",
        operation_id="graph-stage-1",
        component_input=_graph_input(snapshot, operation_id="graph-stage-1"),
    )
    stage = first.stage_public_graph_component(
        grant,
        graph_resource_manifest_hash="graph_manifest_1",
        graph_resource_ids=["graph_resource_1"],
        build_receipt_hash="build_receipt_1",
    )

    restarted = GraphComponentCoordinator(
        SqlAlchemyGenerationManager(SqlAlchemyIndexingRepository(engine)),
        source,
        consumer_id="indexer_1",
    )
    assert (
        restarted.reserve_graph_component_stage(
            graph_build_id="graph_1",
            expected_source_revision=snapshot.source_revision,
            expected_active_generation_id="generation_initial",
            operation_id="graph-stage-1",
            component_input=_graph_input(snapshot, operation_id="graph-stage-1"),
        )
        == grant
    )
    active = restarted.release_graph_component(
        target_generation_id=stage.target_generation_id,
        target_generation_fence=stage.target_generation_fence,
        component_stage_id=stage.component_stage_id,
        source_revision=stage.source_revision,
        source_manifest_hash=stage.source_manifest_hash,
        source_head_fence=stage.source_head_fence,
        operation_id="graph-release-1",
    )
    replay = GraphComponentCoordinator(
        SqlAlchemyGenerationManager(SqlAlchemyIndexingRepository(engine)),
        source,
        consumer_id="indexer_1",
    ).release_graph_component(
        target_generation_id=stage.target_generation_id,
        target_generation_fence=stage.target_generation_fence,
        component_stage_id=stage.component_stage_id,
        source_revision=stage.source_revision,
        source_manifest_hash=stage.source_manifest_hash,
        source_head_fence=stage.source_head_fence,
        operation_id="graph-release-1",
    )

    assert active == replay
    assert active.active_generation_id == stage.target_generation_id


def test_discard_retracts_a_published_attempt() -> None:
    writer = InMemoryIndexWriter()
    writer.stage_chunks(
        "attempt_1",
        "publication_1",
        "document_1",
        "version_1",
        (_chunk("chunk_1"),),
    )
    writer.publish_staged("attempt_1", "publication_1")

    writer.discard_staged("attempt_1", "publication_1")

    assert writer.search("text", ("space_1",), 10, None).items == ()


def test_published_replay_keeps_provider_identity_checks() -> None:
    writer = InMemoryIndexWriter()
    staged = writer.stage_chunks(
        "attempt_1",
        "publication_1",
        "document_1",
        "version_1",
        (_chunk("chunk_1"),),
        fencing_token=7,
        expected_generation_id="generation_initial",
        content_hash="manifest_hash_1",
    )
    writer.publish_staged(
        "attempt_1",
        "publication_1",
        fencing_token=7,
        expected_generation_id="generation_initial",
        content_hash=staged.content_hash,
    )
    with pytest.raises(PlatformError) as error:
        writer.publish_staged(
            "attempt_1",
            "publication_1",
            fencing_token=8,
            expected_generation_id="generation_initial",
            content_hash=staged.content_hash,
        )
    assert error.value.code == "fence_conflict"


class _FailingPublishWriter(InMemoryIndexWriter):
    def publish_staged(self, attempt_id: str, publication_id: str, **kwargs: object):
        del attempt_id, publication_id, kwargs
        raise PlatformError("indexing_publish_failed", "sparse provider failed", {}, 409)


def test_publish_failure_retracts_an_already_published_component() -> None:
    dense = InMemoryIndexWriter(provider_name="dense")
    sparse = _FailingPublishWriter(provider_name="sparse")
    service = IndexingService(dense_writer=dense, sparse_provider=sparse)
    request = _request()
    output = service.process_and_stage(
        request,
        "A heading\n\nbody",
        media_kind="text/plain",
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
    )

    with pytest.raises(PlatformError):
        service.publish(request, receipt=output.receipt)

    assert dense.search("body", ("space_1",), 10, None).items == ()


def test_malformed_receipt_discards_the_actual_staged_resources() -> None:
    dense = InMemoryIndexWriter(provider_name="dense")
    sparse = InMemorySparseIndexProvider()
    service = IndexingService(dense_writer=dense, sparse_provider=sparse)
    request = _request()
    output = service.process_and_stage(
        request,
        "# Heading\n\nretrievable text",
        media_kind="text/markdown",
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
    )
    malformed = output.receipt.to_mapping()
    del malformed["job_id"]

    with pytest.raises(PlatformError) as error:
        service.publish(request, receipt=malformed)

    assert error.value.code == "validation_error"
    assert dense.search("text", ("space_1",), 10, None).items == ()
    assert sparse.search("text", ("space_1",), 10, None).items == ()


class _CapturingPublishWriter(InMemoryIndexWriter):
    def __init__(self, *, provider_name: str) -> None:
        super().__init__(provider_name=provider_name)
        self.publish_kwargs: dict[str, object] | None = None

    def publish_staged(self, attempt_id: str, publication_id: str, **kwargs: object):
        self.publish_kwargs = dict(kwargs)
        return super().publish_staged(attempt_id, publication_id, **kwargs)


class _CapturingStageWriter(InMemoryIndexWriter):
    def __init__(self, *, provider_name: str) -> None:
        super().__init__(provider_name=provider_name)
        self.stage_kwargs: dict[str, object] | None = None

    def stage_chunks(self, attempt_id: str, publication_id: str, *args: object, **kwargs: object):
        self.stage_kwargs = dict(kwargs)
        return super().stage_chunks(attempt_id, publication_id, *args, **kwargs)


def test_publish_binds_provider_call_to_staged_request_identity() -> None:
    dense = _CapturingPublishWriter(provider_name="dense")
    sparse = _CapturingPublishWriter(provider_name="sparse")
    service = IndexingService(dense_writer=dense, sparse_provider=sparse)
    request = _request()
    output = service.process_and_stage(
        request,
        "# Heading\n\nretrievable text",
        media_kind="text/markdown",
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
    )

    service.publish(request, receipt=output.receipt)

    for provider in (dense, sparse):
        assert provider.publish_kwargs is not None
        assert provider.publish_kwargs["fencing_token"] == request.fencing_token
        assert provider.publish_kwargs["expected_generation_id"] == request.expected_generation_id
        assert provider.publish_kwargs["content_hash"] == request.input_manifest_hash
        assert provider.publish_kwargs["stage_resource_manifest"] == output.receipt.stage_resources


def test_process_stages_claimed_document_usage_context() -> None:
    dense = _CapturingStageWriter(provider_name="dense")
    sparse = InMemorySparseIndexProvider()
    service = IndexingService(dense_writer=dense, sparse_provider=sparse)
    request = replace(
        _request(),
        space_id="department:dept_1",
        usage_ownership={
            "actor_user_id": "user_1",
            "actor_role_snapshot": "user",
            "actor_department_id_snapshot": None,
            "quota_subject_user_id": "user_1",
            "cost_center_key": "department:dept_1",
            "space_id": "department:dept_1",
            "space_kind": "department",
            "space_owner_user_id": None,
            "authorization_version": None,
            "fence_token": 1,
            "source_space_ids": ["department:dept_1"],
        },
        usage_deadline_at_utc=datetime(2026, 8, 21, 12, 5, tzinfo=UTC),
        usage_replay_generation=2,
    )

    service.process_and_stage(
        request,
        "# Heading\n\nretrievable text",
        media_kind="text/markdown",
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
    )

    assert dense.stage_kwargs is not None
    context = dense.stage_kwargs["usage_context"]
    assert context.execution_kind == "ingestion"
    assert context.execution_id == request.job_id
    assert context.attempt_id == request.attempt_id
    assert context.generation_id == request.expected_generation_id
    assert context.replay_generation == 2
    assert context.ownership.cost_center_key == "department:dept_1"
    assert context.ownership.source_space_ids == ("department:dept_1",)


def test_staging_request_rejects_boolean_usage_replay_generation() -> None:
    request = _request().to_mapping()
    request["usage_replay_generation"] = True

    with pytest.raises(PlatformError) as error:
        IndexStagingRequest.from_mapping(request)

    assert error.value.code == "validation_error"


def test_retrieval_release_resolves_only_for_its_generation() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    documents_metadata.create_all(engine)
    indexing_metadata.create_all(engine)
    SqlAlchemyIndexingRepository(engine).active_generation_id()
    releases = RetrievalReleaseService(engine)
    staged = releases.stage(
        generation_id="generation_initial",
        profile=RetrievalProfile(),
        acceptance_suite=_release_suite(),
    )
    releases.release(str(staged["id"]), metrics=_release_metrics())

    with pytest.raises(PlatformError) as error:
        releases.resolve(RetrievalProfile(), generation_id="generation_next")

    assert error.value.code == "retrieval_release_unavailable"


def test_retrieval_release_rejects_metrics_outside_the_staged_acceptance_suite() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    documents_metadata.create_all(engine)
    indexing_metadata.create_all(engine)
    SqlAlchemyIndexingRepository(engine).active_generation_id()
    releases = RetrievalReleaseService(engine)
    staged = releases.stage(
        generation_id="generation_initial",
        profile=RetrievalProfile(),
        acceptance_suite=_release_suite(),
    )

    with pytest.raises(PlatformError) as error:
        releases.release(
            str(staged["id"]),
            metrics={**_release_metrics(), "p50_ms": 11},
        )

    assert error.value.code == "release_gate_failed"


def test_generation_release_rejects_orphan_chunk_publications() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    documents_metadata.create_all(engine)
    indexing_metadata.create_all(engine)
    repository = SqlAlchemyIndexingRepository(engine)
    staging = repository.create_staging([], base_revision=0, generation_id="generation_next")
    with engine.begin() as connection:
        repository.record_published_chunks(
            _request(generation=staging.generation_id),
            (_chunk("chunk_orphan", generation=staging.generation_id),),
            connection=connection,
        )

    with pytest.raises(PlatformError) as error:
        repository.release(staging.generation_id)

    assert error.value.code == "release_gate_failed"


def test_processing_receipt_carries_identity_and_table_facts() -> None:
    output = ContentProcessor().process(
        _request(),
        "name,amount\na,1\nb,2\n",
        media_kind="text/csv",
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
    )
    assert output.receipt.processing_summary["table_count"] == 1
    assert output.receipt.stage_resource_ids
    assert output.receipt.input_manifest_hash == "manifest_hash_1"
    assert output.receipt.model_versions["primary"] == "none"
    assert output.receipt.prompt_versions["primary"] == "none"
    output.receipt.validate_against(_request())


@pytest.mark.parametrize("field", ("input_manifest_hash", "processing_profile_version"))
def test_receipt_mapping_requires_stage_identity_echoes(field: str) -> None:
    receipt = (
        ContentProcessor()
        .process(
            _request(),
            "# Heading\n\nbody",
            media_kind="text/markdown",
            content_manifest_id="manifest_1",
            content_manifest_hash="manifest_hash_1",
        )
        .receipt.to_mapping()
    )
    del receipt[field]

    with pytest.raises(PlatformError) as error:
        IndexProcessingReceipt.from_mapping(receipt)

    assert error.value.code == "validation_error"


def test_excel_uses_structured_table_loader() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Revenue"
    worksheet.append(["month", "amount"])
    worksheet.append(["January", 10])
    stream = io.BytesIO()
    workbook.save(stream)
    output = ContentProcessor().process(
        _request(),
        stream.getvalue(),
        media_kind="xlsx",
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
    )
    assert output.chunks[0].locator == {"sheet": "Revenue", "a1_range": "A2:B2"}
    assert output.receipt.processing_summary["sheet_count"] == 1


def test_header_only_csv_preview_metadata_matches_content_rows() -> None:
    source = b"name\n"
    output = ContentProcessor().process(
        _request(),
        source,
        media_kind="text/csv",
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
    )
    renderer = ProcessingReceiptPreviewRenderer()
    metadata = renderer.metadata(processing_summary=output.receipt.processing_summary)
    content = renderer.render(
        version={"media_kind": "text/csv", "processing_summary": output.receipt.processing_summary},
        content=source,
        metadata=metadata,
        sheet="CSV",
    )

    assert metadata.sheets == ({"name": "CSV", "row_count": 1},)
    assert json.loads(content.body)["row_count"] == 1


def test_header_only_xlsx_preview_metadata_matches_content_rows() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["name"])
    stream = io.BytesIO()
    workbook.save(stream)
    source = stream.getvalue()
    output = ContentProcessor().process(
        _request(),
        source,
        media_kind="xlsx",
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
    )
    renderer = ProcessingReceiptPreviewRenderer()
    metadata = renderer.metadata(processing_summary=output.receipt.processing_summary)
    content = renderer.render(
        version={"media_kind": "xlsx", "processing_summary": output.receipt.processing_summary},
        content=source,
        metadata=metadata,
        sheet="Sheet",
    )

    assert metadata.sheets == ({"name": "Sheet", "row_count": 1},)
    assert json.loads(content.body)["row_count"] == 1


def test_excel_expands_merged_cells_and_records_ranges() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Revenue"
    worksheet.append(["category", "amount"])
    worksheet.append(["North", 10])
    worksheet.append([None, 20])
    worksheet.append(["South", 30])
    worksheet.append([None, 40])
    worksheet.merge_cells("A2:A3")
    worksheet.merge_cells("A4:A5")
    stream = io.BytesIO()
    workbook.save(stream)

    output = ContentProcessor().process(
        _request(),
        stream.getvalue(),
        media_kind="xlsx",
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
    )

    assert "North" in output.chunks[0].text
    assert "South" in output.chunks[0].text
    assert set(output.receipt.processing_summary["sheet_manifest"][0]["merged_ranges"]) == {
        "A2:A3",
        "A4:A5",
    }


def test_excel_streams_merged_range_scan_without_materializing_worksheet_xml(monkeypatch) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Revenue"
    worksheet.append(["category", "amount"])
    worksheet.append(["North", 10])
    worksheet.append([None, 20])
    worksheet.append(["South", 30])
    worksheet.append([None, 40])
    worksheet.merge_cells("A2:A3")
    worksheet.merge_cells("A4:A5")
    stream = io.BytesIO()
    workbook.save(stream)

    original_read = zipfile.ZipFile.read

    def reject_worksheet_read(archive, name, *args, **kwargs):
        filename = getattr(name, "filename", name)
        if str(filename).startswith("xl/worksheets/"):
            raise AssertionError("worksheet XML must be scanned as a stream")
        return original_read(archive, name, *args, **kwargs)

    monkeypatch.setattr(zipfile.ZipFile, "read", reject_worksheet_read)

    assert set(_xlsx_merged_ranges(stream.getvalue(), maximum_cells=10_000)["Revenue"]) == {
        "A2:A3",
        "A4:A5",
    }


def test_unheaded_text_is_split_into_bounded_chunks() -> None:
    output = ContentProcessor().process(
        _request(),
        "x" * 16_001,
        media_kind="text/plain",
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
    )

    assert len(output.chunks) == 3
    assert all(len(chunk.text) <= 8_000 for chunk in output.chunks)
    assert "".join(chunk.text for chunk in output.chunks) == "x" * 16_001


@pytest.mark.parametrize(
    ("content", "media_kind"),
    (
        ("value\n" + "x" * 9, "text/csv"),
        ("x" * 9, "text/x-python"),
    ),
)
def test_csv_and_code_chunks_respect_configured_character_limit(
    content: str, media_kind: str
) -> None:
    output = ContentProcessor(text_chunk_max_chars=7).process(
        _request(),
        content,
        media_kind=media_kind,
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
    )

    assert output.chunks
    assert all(len(chunk.text) <= 7 for chunk in output.chunks)


@pytest.mark.parametrize(
    ("content", "media_kind"),
    (("plain text", "text/plain"), ("name,amount\na,1\n", "text/csv")),
)
def test_nonpaged_content_is_billed_as_one_page(content: str, media_kind: str) -> None:
    output = ContentProcessor().process(
        _request(),
        content,
        media_kind=media_kind,
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
    )

    assert output.receipt.processing_summary["page_count"] == 1


def test_xml_entity_declarations_are_rejected() -> None:
    with pytest.raises(PlatformError) as error:
        ContentProcessor().process(
            _request(),
            "<!DOCTYPE root [<!ENTITY expanded 'value'>]><root>&expanded;</root>",
            media_kind="application/xml",
            content_manifest_id="manifest_1",
            content_manifest_hash="manifest_hash_1",
        )

    assert error.value.code == "structured_parse_failed"


def test_xml_without_entity_declarations_remains_processible() -> None:
    output = ContentProcessor().process(
        _request(),
        "<!DOCTYPE root><root><item>ok</item></root>",
        media_kind="application/xml",
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
    )

    assert output.chunks


def test_excel_rejects_a_merged_region_above_the_supported_limit() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "merged"
    worksheet.merge_cells("A1:A10001")
    stream = io.BytesIO()
    workbook.save(stream)

    with pytest.raises(PlatformError) as error:
        ContentProcessor().process(
            _request(),
            stream.getvalue(),
            media_kind="xlsx",
            content_manifest_id="manifest_1",
            content_manifest_hash="manifest_hash_1",
        )

    assert error.value.code == "table_parse_failed"


def test_mineru_pdf_preserves_native_page_and_span_locators() -> None:
    processor = ContentProcessor(
        mineru=lambda content: {
            "text": "# First\nfirst\n\n# Second\nsecond",
            "page_count": 2,
            "has_text_layer": True,
            "chunks": [{"page": 1, "span": "0:5"}, {"page": 2, "span": "0:6"}],
        }
    )

    output = processor.process(
        _request(),
        b"pdf",
        media_kind="application/pdf",
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
    )

    assert output.chunks[0].locator == {"page": 1, "span": "0:5"}
    assert output.chunks[1].locator == {"page": 2, "span": "0:6"}
    assert all(chunk.snippet for chunk in output.chunks)
    assert output.receipt.processing_summary["page_count"] == 2


def test_standard_docx_mime_is_accepted_and_routes_through_mineru() -> None:
    media_kind = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    upload = DocumentUpload(filename="guide.docx", content=b"word", media_kind=media_kind)
    processed: list[bytes] = []
    processor = ContentProcessor(
        mineru=lambda content: (
            processed.append(content)
            or {
                "text": "# Guide\n\nWord content",
                "page_count": 1,
                "has_text_layer": True,
                "chunks": [{"page": 1, "span": "0:12"}],
            }
        )
    )

    output = processor.process(
        _request(),
        upload.content,
        media_kind=upload.media_kind,
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
    )

    assert processed == [b"word"]
    assert output.chunks[0].locator == {"section_path": "Guide"}
    assert output.chunks[0].snippet is None
    assert output.receipt.processing_summary["tree"]["sections"] == [
        {"path": ["Guide"], "paragraphs": ["Word content"]}
    ]


def test_processor_word_tree_receipt_renders_json_without_decoding_source() -> None:
    output = ContentProcessor(
        mineru=lambda content: {"text": "# Guide\n\nWord content", "page_count": 1}
    ).process(
        _request(),
        b"\xff\xfe\x80",
        media_kind="docx",
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
    )
    renderer = ProcessingReceiptPreviewRenderer()
    metadata = renderer.metadata(processing_summary=output.receipt.processing_summary)

    content = renderer.render(
        version={"media_kind": "docx", "processing_summary": output.receipt.processing_summary},
        content=b"\xff\xfe\x80",
        metadata=metadata,
        sheet=None,
    )

    assert metadata.tree_indexed is True
    assert content.media_type == "application/json"
    assert json.loads(content.body) == {
        "sections": [{"path": ["Guide"], "paragraphs": ["Word content"]}]
    }


def test_processor_basic_word_receipt_renders_readable_text_without_decoding_source() -> None:
    output = ContentProcessor(
        mineru=lambda content: {"text": "Readable plain text", "page_count": 1}
    ).process(
        _request(),
        b"\xff\xfe\x80",
        media_kind="docx",
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
    )
    renderer = ProcessingReceiptPreviewRenderer()
    metadata = renderer.metadata(processing_summary=output.receipt.processing_summary)

    content = renderer.render(
        version={"media_kind": "docx", "processing_summary": output.receipt.processing_summary},
        content=b"\xff\xfe\x80",
        metadata=metadata,
        sheet=None,
    )

    assert metadata.tree_indexed is False
    assert content.media_type == "text/plain"
    assert content.body == b"Readable plain text"


def test_structured_large_configuration_calls_compressor() -> None:
    calls: list[dict[str, object]] = []

    class Compressor:
        def compress(self, text: str, *, context: dict[str, object]) -> str:
            calls.append({"text": text, **context})
            return "compressed config"

    payload = "{" + ",".join(f'"key_{index}": "value"' for index in range(600)) + "}"
    output = ContentProcessor(compressor=Compressor()).process(
        _request(),
        payload,
        media_kind="json",
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
    )

    assert calls
    assert output.chunks[0].embedding_text == "compressed config"
    assert output.receipt.processing_summary["cr"]["applied"] is True


def test_image_description_receives_context_and_ocr_text() -> None:
    seen: dict[str, object] = {}

    def describe(content: bytes, context: dict[str, object]) -> str:
        seen.update(context)
        return "diagram description"

    output = ContentProcessor(image_describer=describe).process(
        _request(),
        b"image",
        media_kind="image/png",
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
        image_context={"caption": "Revenue chart", "section_path": "2 / Results"},
        image_ocr_text="Q1: 10",
    )

    assert seen["caption"] == "Revenue chart"
    assert output.chunks[0].text.endswith("Q1: 10")


def test_image_pipeline_runs_ocr_and_keeps_caption_and_reverse_links_with_description() -> None:
    seen: dict[str, object] = {}

    def describe(content: bytes, context: dict[str, object]) -> str:
        del content
        seen.update(context)
        return "diagram description"

    def ocr(content: bytes, context: dict[str, object]) -> str:
        assert content == b"image"
        assert context["section_path"] == "2 / Results"
        return "Q1: 10"

    output = ContentProcessor(image_describer=describe, image_ocr=ocr).process(
        _request(),
        b"image",
        media_kind="image/png",
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
        image_context={
            "caption": "Revenue chart",
            "preceding_text": "Revenue rose in the quarter.",
            "following_text": "The chart summarizes the result.",
            "section_path": "2 / Results",
            "reverse_links": ["chunk_overview"],
        },
    )

    assert seen["preceding_text"] == "Revenue rose in the quarter."
    assert output.chunks[0].text == "Revenue chart\ndiagram description\nQ1: 10"
    assert output.chunks[0].metadata["reverse_links"] == ["chunk_overview"]
    assert output.receipt.processing_summary["ocr"]["applied"] is True


def test_indexing_service_stages_then_publishes_without_document_fact_writes() -> None:
    service = IndexingService()
    request = _request()
    output = service.process_and_stage(
        request,
        "# Heading\n\nretrievable text",
        media_kind="text/markdown",
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
    )
    published = service.publish(request, receipt=output.receipt)
    assert published["state"] == "published"


def test_processing_rejects_a_content_manifest_that_is_not_the_staging_input() -> None:
    with pytest.raises(PlatformError) as error:
        ContentProcessor().process(
            _request(),
            "content",
            media_kind="text/plain",
            content_manifest_id="manifest_1",
            content_manifest_hash="different_manifest_hash",
        )

    assert error.value.code == "processing_receipt_conflict"


def test_image_without_usable_text_is_published_as_non_indexable() -> None:
    service = IndexingService()
    request = _request()
    output = service.process_and_stage(
        request,
        b"image-bytes",
        media_kind="image/png",
        content_manifest_id="manifest_1",
        content_manifest_hash="manifest_hash_1",
    )
    assert output.chunks == ()
    assert output.receipt.processing_summary["image_count"] == 1
    assert service.publish(request, receipt=output.receipt)["state"] == "published"


def test_stale_graph_source_cannot_activate_a_generation() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    documents_metadata.create_all(engine)
    source = PublicGraphSourceService(
        engine,
        trusted_consumers={"indexing": {"indexer_1"}},
        outbox_port=_Outbox(),
    )
    first = source.record_source_change(
        space_id="public",
        document_id="document_1",
        change_type="publish",
        publications=[],
    )
    manager = GenerationManager()
    coordinator = GraphComponentCoordinator(manager, source, consumer_id="indexer_1")
    grant = coordinator.reserve_graph_component_stage(
        graph_build_id="graph_1",
        expected_source_revision=first.source_revision,
        expected_active_generation_id=manager.active_generation_id,
        operation_id="stage_1",
        component_input=_graph_input(first, operation_id="stage_1"),
    )
    receipt = coordinator.stage_public_graph_component(
        grant,
        graph_resource_manifest_hash="graph_manifest_1",
        graph_resource_ids=["graph_resource_1"],
        build_receipt_hash="build_receipt_1",
    )
    source.record_source_change(
        space_id="public",
        document_id="document_2",
        change_type="publish",
        publications=[],
    )
    with pytest.raises(PlatformError) as error:
        coordinator.release_graph_component(
            target_generation_id=grant.target_generation_id,
            target_generation_fence=grant.target_generation_fence,
            component_stage_id=receipt.component_stage_id,
            source_revision=receipt.source_revision,
            source_manifest_hash=receipt.source_manifest_hash,
            source_head_fence=receipt.source_head_fence,
            operation_id="release_1",
        )
    assert error.value.code == "graph_source_changed"
    assert manager.active_generation_id == "generation_initial"


def test_graph_reader_lease_rechecks_current_source_head() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    documents_metadata.create_all(engine)
    source = PublicGraphSourceService(
        engine,
        trusted_consumers={"indexing": {"indexer_1"}},
        outbox_port=_Outbox(),
    )
    first = source.record_source_change(
        space_id="public",
        document_id="document_1",
        change_type="publish",
        publications=[],
    )
    manager = GenerationManager()
    coordinator = GraphComponentCoordinator(manager, source, consumer_id="indexer_1")
    grant = coordinator.reserve_graph_component_stage(
        graph_build_id="graph_1",
        expected_source_revision=first.source_revision,
        expected_active_generation_id=manager.active_generation_id,
        operation_id="stage_1",
        component_input=_graph_input(first, operation_id="stage_1"),
    )
    stage = coordinator.stage_public_graph_component(
        grant,
        graph_resource_manifest_hash="graph_manifest_1",
        graph_resource_ids=["graph_resource_1"],
        build_receipt_hash="build_receipt_1",
    )
    coordinator.release_graph_component(
        target_generation_id=grant.target_generation_id,
        target_generation_fence=grant.target_generation_fence,
        component_stage_id=stage.component_stage_id,
        source_revision=stage.source_revision,
        source_manifest_hash=stage.source_manifest_hash,
        source_head_fence=stage.source_head_fence,
        operation_id="release_1",
    )
    lease = coordinator.acquire_reader_lease(
        source_revision=stage.source_revision,
        source_manifest_hash=stage.source_manifest_hash,
        source_head_fence=stage.source_head_fence,
        manifest_hash="graph_manifest_1",
    )
    second = source.record_source_change(
        space_id="public",
        document_id="document_2",
        change_type="publish",
        publications=[],
    )
    second_head = source.get_current_head()
    with pytest.raises(PlatformError) as error:
        coordinator.renew_reader_lease(
            lease.lease_id,
            source_revision=second.source_revision,
            source_manifest_hash=second.source_manifest_hash,
            source_head_fence=second_head.source_head_fence,
        )
    assert error.value.code == "graph_source_changed"
    stale = manager.get_generation(manager.active_generation_id)
    assert stale.graph_component_state == "stale"
    assert stale.manifest["components"]["public_graph"]["graph_resource_ids"] == []
    assert lease.lease_id not in manager._leases


def test_invalid_graph_release_receipt_discards_staged_graph_resources() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    documents_metadata.create_all(engine)
    indexing_metadata.create_all(engine)
    source = PublicGraphSourceService(
        engine,
        trusted_consumers={"indexing": {"indexer_1"}},
        outbox_port=_Outbox(),
    )
    snapshot = source.record_source_change(
        space_id="public", document_id="document_1", change_type="publish", publications=[]
    )
    manager = SqlAlchemyGenerationManager(SqlAlchemyIndexingRepository(engine))
    coordinator = GraphComponentCoordinator(manager, source, consumer_id="indexer_1")
    grant = coordinator.reserve_graph_component_stage(
        graph_build_id="graph_1",
        expected_source_revision=snapshot.source_revision,
        expected_active_generation_id=manager.active_generation_id,
        operation_id="stage_1",
        component_input=_graph_input(snapshot, operation_id="stage_1"),
    )
    receipt = coordinator.stage_public_graph_component(
        grant,
        graph_resource_manifest_hash="graph_manifest_1",
        graph_resource_ids=["graph_resource_1"],
        build_receipt_hash="build_receipt_1",
    )

    with pytest.raises(PlatformError) as error:
        coordinator.release_graph_component(
            target_generation_id=grant.target_generation_id,
            target_generation_fence="wrong_fence",
            component_stage_id=receipt.component_stage_id,
            source_revision=receipt.source_revision,
            source_manifest_hash=receipt.source_manifest_hash,
            source_head_fence=receipt.source_head_fence,
            operation_id="release_1",
        )

    assert error.value.code == "processing_receipt_conflict"
    staged = manager.get_generation(grant.target_generation_id)
    assert staged.status == "staging"
    assert staged.manifest["components"]["public_graph"]["state"] == "disabled"
    assert staged.manifest["components"]["public_graph"]["graph_resource_ids"] == []
    assert manager.active_generation_id == "generation_initial"


def test_unknown_graph_stage_id_discards_authoritative_staged_resources() -> None:
    class RecordingSource(PublicGraphSourceService):
        def __init__(self, *args, **kwargs) -> None:
            self.acknowledgements: list[str] = []
            super().__init__(*args, **kwargs)

        def acknowledge_consumption(self, **kwargs):
            self.acknowledgements.append(str(kwargs["purpose"]))
            return super().acknowledge_consumption(**kwargs)

    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    documents_metadata.create_all(engine)
    indexing_metadata.create_all(engine)
    source = RecordingSource(
        engine,
        trusted_consumers={"indexing": {"indexer_1"}},
        outbox_port=_Outbox(),
    )
    snapshot = source.record_source_change(
        space_id="public", document_id="document_1", change_type="publish", publications=[]
    )
    manager = SqlAlchemyGenerationManager(SqlAlchemyIndexingRepository(engine))
    coordinator = GraphComponentCoordinator(manager, source, consumer_id="indexer_1")
    grant = coordinator.reserve_graph_component_stage(
        graph_build_id="graph_1",
        expected_source_revision=snapshot.source_revision,
        expected_active_generation_id=manager.active_generation_id,
        operation_id="stage_1",
        component_input=_graph_input(snapshot, operation_id="stage_1"),
    )
    coordinator.stage_public_graph_component(
        grant,
        graph_resource_manifest_hash="graph_manifest_1",
        graph_resource_ids=["graph_resource_1"],
        build_receipt_hash="build_receipt_1",
    )

    with pytest.raises(PlatformError) as error:
        coordinator.release_graph_component(
            target_generation_id=grant.target_generation_id,
            target_generation_fence=grant.target_generation_fence,
            component_stage_id="unknown_stage",
            source_revision=snapshot.source_revision,
            source_manifest_hash=snapshot.source_manifest_hash,
            source_head_fence=source.get_current_head().source_head_fence,
            operation_id="release_1",
        )

    assert error.value.code == "processing_receipt_conflict"
    graph = manager.get_generation(grant.target_generation_id).manifest["components"][
        "public_graph"
    ]
    assert graph["state"] == "disabled"
    assert graph["graph_resource_ids"] == []
    assert source.acknowledgements == ["stage", "discard"]
    assert manager.active_generation_id == "generation_initial"


def test_rollback_disables_ready_graph_without_authoritative_source_validation() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    documents_metadata.create_all(engine)
    indexing_metadata.create_all(engine)
    repository = SqlAlchemyIndexingRepository(engine)
    repository.set_component_state(
        "generation_initial",
        "public_graph",
        "ready",
        manifest={
            "source_revision": 1,
            "source_manifest_hash": "source_manifest_1",
            "source_head_fence": 1,
            "graph_resource_ids": ["graph_resource_1"],
        },
    )
    staging = repository.create_staging([], generation_id="generation_next")
    repository.release(staging.generation_id)

    restored = repository.rollback(
        "generation_initial",
        source_receipt={
            "state": "held",
            "candidate_generation_id": "generation_initial",
            "applied_revision": 0,
            "source_revision": 1,
            "source_manifest_hash": "source_manifest_1",
            "source_head_fence": 1,
        },
    )

    assert restored.generation_id == "generation_initial"
    assert restored.manifest["components"]["public_graph"]["state"] == "disabled"
    assert restored.manifest["components"]["public_graph"]["graph_resource_ids"] == []


def test_graph_rollback_retains_ready_component_only_after_current_source_validation() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    documents_metadata.create_all(engine)
    indexing_metadata.create_all(engine)
    source = PublicGraphSourceService(
        engine,
        trusted_consumers={"indexing": {"indexer_1"}},
        outbox_port=_Outbox(),
    )
    snapshot = source.record_source_change(
        space_id="public", document_id="document_1", change_type="publish", publications=[]
    )
    repository = SqlAlchemyIndexingRepository(engine)
    manager = SqlAlchemyGenerationManager(repository)
    coordinator = GraphComponentCoordinator(manager, source, consumer_id="indexer_1")
    grant = coordinator.reserve_graph_component_stage(
        graph_build_id="graph_1",
        expected_source_revision=snapshot.source_revision,
        expected_active_generation_id=manager.active_generation_id,
        operation_id="stage_1",
        component_input=_graph_input(snapshot, operation_id="stage_1"),
    )
    stage = coordinator.stage_public_graph_component(
        grant,
        graph_resource_manifest_hash="graph_manifest_1",
        graph_resource_ids=["graph_resource_1"],
        build_receipt_hash="build_receipt_1",
    )
    coordinator.release_graph_component(
        target_generation_id=grant.target_generation_id,
        target_generation_fence=grant.target_generation_fence,
        component_stage_id=stage.component_stage_id,
        source_revision=stage.source_revision,
        source_manifest_hash=stage.source_manifest_hash,
        source_head_fence=stage.source_head_fence,
        operation_id="release_1",
    )
    final = repository.create_staging([], generation_id="generation_final")
    repository.release(final.generation_id)

    restored = coordinator.rollback_generation(
        candidate_generation_id=grant.target_generation_id,
        source_receipt={
            "state": "held",
            "candidate_generation_id": grant.target_generation_id,
            "applied_revision": 0,
            "source_revision": stage.source_revision,
            "source_manifest_hash": stage.source_manifest_hash,
            "source_head_fence": stage.source_head_fence,
        },
        operation_id="rollback_1",
        caller_principal="ops",
    )

    assert restored.generation_id == grant.target_generation_id
    graph = restored.manifest["components"]["public_graph"]
    assert graph["state"] == "ready"
    assert graph["graph_resource_ids"] == ["graph_resource_1"]


def test_graph_rollback_disables_stale_ready_component_before_activation() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    documents_metadata.create_all(engine)
    indexing_metadata.create_all(engine)
    source = PublicGraphSourceService(
        engine,
        trusted_consumers={"indexing": {"indexer_1"}},
        outbox_port=_Outbox(),
    )
    snapshot = source.record_source_change(
        space_id="public", document_id="document_1", change_type="publish", publications=[]
    )
    repository = SqlAlchemyIndexingRepository(engine)
    manager = SqlAlchemyGenerationManager(repository)
    coordinator = GraphComponentCoordinator(manager, source, consumer_id="indexer_1")
    grant = coordinator.reserve_graph_component_stage(
        graph_build_id="graph_1",
        expected_source_revision=snapshot.source_revision,
        expected_active_generation_id=manager.active_generation_id,
        operation_id="stage_1",
        component_input=_graph_input(snapshot, operation_id="stage_1"),
    )
    stage = coordinator.stage_public_graph_component(
        grant,
        graph_resource_manifest_hash="graph_manifest_1",
        graph_resource_ids=["graph_resource_1"],
        build_receipt_hash="build_receipt_1",
    )
    coordinator.release_graph_component(
        target_generation_id=grant.target_generation_id,
        target_generation_fence=grant.target_generation_fence,
        component_stage_id=stage.component_stage_id,
        source_revision=stage.source_revision,
        source_manifest_hash=stage.source_manifest_hash,
        source_head_fence=stage.source_head_fence,
        operation_id="release_1",
    )
    final = repository.create_staging([], generation_id="generation_final")
    repository.release(final.generation_id)
    source.record_source_change(
        space_id="public", document_id="document_2", change_type="publish", publications=[]
    )

    restored = coordinator.rollback_generation(
        candidate_generation_id=grant.target_generation_id,
        source_receipt={
            "state": "held",
            "candidate_generation_id": grant.target_generation_id,
            "applied_revision": 0,
            "source_revision": stage.source_revision,
            "source_manifest_hash": stage.source_manifest_hash,
            "source_head_fence": stage.source_head_fence,
        },
        operation_id="rollback_1",
        caller_principal="ops",
    )

    assert restored.generation_id == grant.target_generation_id
    graph = restored.manifest["components"]["public_graph"]
    assert graph["state"] == "disabled"
    assert graph["graph_resource_ids"] == []
    assert manager.active_generation_id == grant.target_generation_id


def test_gc_completion_rechecks_rollback_blocker_after_retention_request() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    documents_metadata.create_all(engine)
    indexing_metadata.create_all(engine)
    repository = SqlAlchemyIndexingRepository(engine)
    first = repository.create_staging([], generation_id="generation_first")
    repository.release(first.generation_id)
    second = repository.create_staging([], generation_id="generation_second")
    repository.release(second.generation_id)
    request = repository.request_index_generation_gc(
        "generation_initial", reconciliation_run_id="reconcile_1", operation_id="gc_1"
    )
    assert request.state == "accepted"

    with engine.begin() as connection:
        connection.execute(
            update(index_generation_heads_table)
            .where(index_generation_heads_table.c.id == "instance")
            .values(rollback_candidate_id="generation_initial")
        )
    completed = repository.complete_generation_gc("generation_initial", operation_id="gc_1")

    assert completed.state == "blocked"
    assert completed.blocking_reasons == ("rollback_candidate",)
    assert repository.get_generation("generation_initial").status == "retired"


def test_gc_completion_rejects_an_operation_for_a_different_generation() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    documents_metadata.create_all(engine)
    indexing_metadata.create_all(engine)
    repository = SqlAlchemyIndexingRepository(engine)
    first = repository.create_staging([], generation_id="generation_first")
    repository.release(first.generation_id)
    second = repository.create_staging([], generation_id="generation_second")
    repository.release(second.generation_id)
    third = repository.create_staging([], generation_id="generation_third")
    repository.release(third.generation_id)
    request = repository.request_index_generation_gc(
        "generation_initial", reconciliation_run_id="reconcile_1", operation_id="gc_initial"
    )
    assert request.state == "accepted"

    with pytest.raises(PlatformError) as error:
        repository.complete_generation_gc("generation_first", operation_id="gc_initial")

    assert error.value.code == "idempotency_key_conflict"
    assert repository.get_generation("generation_first").status == "retired"


def test_generation_gc_purges_only_retired_generation_provider_resources() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    documents_metadata.create_all(engine)
    indexing_metadata.create_all(engine)
    repository = SqlAlchemyIndexingRepository(engine)
    manager = SqlAlchemyGenerationManager(repository)
    dense = InMemoryIndexWriter(provider_name="dense")
    sparse = InMemorySparseIndexProvider()
    IndexingService(
        dense_writer=dense,
        sparse_provider=sparse,
        generation_manager=manager,
        object_store=MemoryObjectStore(),
    )
    retired_generation_id = manager.active_generation_id
    survivor = repository.create_staging([], generation_id="generation_survivor")
    repository.release(survivor.generation_id)
    active = repository.create_staging([], generation_id="generation_active")
    repository.release(active.generation_id)

    for provider in (dense, sparse):
        retired_chunk = _chunk("retired_chunk", generation=retired_generation_id)
        retired_chunk = IndexChunk(
            chunk_id=retired_chunk.chunk_id,
            generation_id=retired_chunk.generation_id,
            publication_id="retired_publication",
            document_id=retired_chunk.document_id,
            document_version_id=retired_chunk.document_version_id,
            space_id=retired_chunk.space_id,
            text=retired_chunk.text,
            embedding_text=retired_chunk.embedding_text,
            locator=retired_chunk.locator,
            snippet=retired_chunk.snippet,
            media_kind=retired_chunk.media_kind,
            manifest_hash=retired_chunk.manifest_hash,
            metadata=retired_chunk.metadata,
            sparse_text=retired_chunk.sparse_text,
            indexable=retired_chunk.indexable,
        )
        provider.stage_chunks(
            "retired_attempt",
            "retired_publication",
            "document_1",
            "version_1",
            [retired_chunk],
            expected_generation_id=retired_generation_id,
            content_hash="retired-content",
        )
        provider.publish_staged(
            "retired_attempt",
            "retired_publication",
            expected_generation_id=retired_generation_id,
            content_hash="retired-content",
        )
        active_chunk = _chunk(
            "active_chunk", generation=active.generation_id, document_id="document_2"
        )
        active_chunk = IndexChunk(
            chunk_id=active_chunk.chunk_id,
            generation_id=active_chunk.generation_id,
            publication_id="active_publication",
            document_id=active_chunk.document_id,
            document_version_id="version_2",
            space_id=active_chunk.space_id,
            text=active_chunk.text,
            embedding_text=active_chunk.embedding_text,
            locator=active_chunk.locator,
            snippet=active_chunk.snippet,
            media_kind=active_chunk.media_kind,
            manifest_hash=active_chunk.manifest_hash,
            metadata=active_chunk.metadata,
            sparse_text=active_chunk.sparse_text,
            indexable=active_chunk.indexable,
        )
        provider.stage_chunks(
            "active_attempt",
            "active_publication",
            "document_2",
            "version_2",
            [active_chunk],
            expected_generation_id=active.generation_id,
            content_hash="active-content",
        )
        provider.publish_staged(
            "active_attempt",
            "active_publication",
            expected_generation_id=active.generation_id,
            content_hash="active-content",
        )

    with engine.begin() as connection:
        repository.record_published_chunks(
            _request(generation=retired_generation_id),
            [_chunk("retired_chunk", generation=retired_generation_id)],
            connection=connection,
        )

    request = repository.request_index_generation_gc(
        retired_generation_id,
        reconciliation_run_id="reconcile_1",
        operation_id="gc_1",
    )
    completed = repository.complete_generation_gc(retired_generation_id, operation_id="gc_1")

    assert request.state == "accepted"
    assert completed.state == "already_purged"
    assert {chunk.generation_id for chunk in dense.visible_chunks()} == {active.generation_id}
    assert {chunk.generation_id for chunk in sparse.visible_chunks()} == {active.generation_id}


def test_graph_reader_source_mismatch_discards_resources_before_rejecting_new_read() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    documents_metadata.create_all(engine)
    source = PublicGraphSourceService(
        engine,
        trusted_consumers={"indexing": {"indexer_1"}},
        outbox_port=_Outbox(),
    )
    first = source.record_source_change(
        space_id="public", document_id="document_1", change_type="publish", publications=[]
    )
    manager = GenerationManager()
    coordinator = GraphComponentCoordinator(manager, source, consumer_id="indexer_1")
    grant = coordinator.reserve_graph_component_stage(
        graph_build_id="graph_1",
        expected_source_revision=first.source_revision,
        expected_active_generation_id=manager.active_generation_id,
        operation_id="stage_1",
        component_input=_graph_input(first, operation_id="stage_1"),
    )
    stage = coordinator.stage_public_graph_component(
        grant,
        graph_resource_manifest_hash="graph_manifest_1",
        graph_resource_ids=["graph_resource_1"],
        build_receipt_hash="build_receipt_1",
    )
    coordinator.release_graph_component(
        target_generation_id=grant.target_generation_id,
        target_generation_fence=grant.target_generation_fence,
        component_stage_id=stage.component_stage_id,
        source_revision=stage.source_revision,
        source_manifest_hash=stage.source_manifest_hash,
        source_head_fence=stage.source_head_fence,
        operation_id="release_1",
    )
    second = source.record_source_change(
        space_id="public", document_id="document_2", change_type="publish", publications=[]
    )
    second_head = source.get_current_head()

    with pytest.raises(PlatformError) as error:
        coordinator.acquire_reader_lease(
            source_revision=second.source_revision,
            source_manifest_hash=second.source_manifest_hash,
            source_head_fence=second_head.source_head_fence,
            manifest_hash="graph_manifest_1",
        )

    assert error.value.code == "graph_source_changed"
    graph = manager.get_generation(manager.active_generation_id).manifest["components"][
        "public_graph"
    ]
    assert graph["state"] == "stale"
    assert graph["graph_resource_ids"] == []


def test_conflicting_graph_stage_receipt_discards_staged_resources() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    documents_metadata.create_all(engine)
    source = PublicGraphSourceService(
        engine,
        trusted_consumers={"indexing": {"indexer_1"}},
        outbox_port=_Outbox(),
    )
    first = source.record_source_change(
        space_id="public", document_id="document_1", change_type="publish", publications=[]
    )
    manager = GenerationManager()
    coordinator = GraphComponentCoordinator(manager, source, consumer_id="indexer_1")
    grant = coordinator.reserve_graph_component_stage(
        graph_build_id="graph_1",
        expected_source_revision=first.source_revision,
        expected_active_generation_id=manager.active_generation_id,
        operation_id="stage_1",
        component_input=_graph_input(first, operation_id="stage_1"),
    )
    coordinator.stage_public_graph_component(
        grant,
        graph_resource_manifest_hash="graph_manifest_1",
        graph_resource_ids=["graph_resource_1"],
        build_receipt_hash="build_receipt_1",
    )

    with pytest.raises(PlatformError) as error:
        coordinator.stage_public_graph_component(
            grant,
            graph_resource_manifest_hash="graph_manifest_2",
            graph_resource_ids=["graph_resource_2"],
            build_receipt_hash="build_receipt_2",
        )

    assert error.value.code == "idempotency_key_conflict"
    graph = manager.get_generation(grant.target_generation_id).manifest["components"][
        "public_graph"
    ]
    assert graph["state"] == "disabled"
    assert graph["graph_resource_ids"] == []


def test_durable_graph_gc_discards_component_before_source_hold_and_chunks() -> None:
    class OrderingSource(PublicGraphSourceService):
        def __init__(self, *args, **kwargs) -> None:
            self.events: list[tuple[str, str, tuple[str, ...], int]] = []
            self.repository: SqlAlchemyIndexingRepository | None = None
            self.candidate_generation_id = ""
            super().__init__(*args, **kwargs)

        def acknowledge_consumption(self, **kwargs):
            if kwargs["purpose"] == "discard":
                assert self.repository is not None
                connection = kwargs["connection"]
                candidate = self.repository.get_generation(
                    self.candidate_generation_id, connection=connection
                )
                graph = candidate.manifest["components"]["public_graph"]
                remaining_chunks = connection.execute(
                    select(index_chunks_table.c.id).where(
                        index_chunks_table.c.generation_id == self.candidate_generation_id
                    )
                ).all()
                self.events.append(
                    (
                        str(graph["state"]),
                        str(graph["graph_resource_manifest_hash"]),
                        tuple(graph["graph_resource_ids"]),
                        len(remaining_chunks),
                    )
                )
            return super().acknowledge_consumption(**kwargs)

    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    documents_metadata.create_all(engine)
    indexing_metadata.create_all(engine)
    source = OrderingSource(
        engine,
        trusted_consumers={"indexing": {"indexer_1"}},
        outbox_port=_Outbox(),
    )
    snapshot = source.record_source_change(
        space_id="public", document_id="document_1", change_type="publish", publications=[]
    )
    repository = SqlAlchemyIndexingRepository(engine)
    source.repository = repository
    manager = SqlAlchemyGenerationManager(repository)
    coordinator = GraphComponentCoordinator(manager, source, consumer_id="indexer_1")
    grant = coordinator.reserve_graph_component_stage(
        graph_build_id="graph_1",
        expected_source_revision=snapshot.source_revision,
        expected_active_generation_id=manager.active_generation_id,
        operation_id="stage_1",
        component_input=_graph_input(snapshot, operation_id="stage_1"),
    )
    stage = coordinator.stage_public_graph_component(
        grant,
        graph_resource_manifest_hash="graph_manifest_1",
        graph_resource_ids=["graph_resource_1"],
        build_receipt_hash="build_receipt_1",
    )
    coordinator.release_graph_component(
        target_generation_id=grant.target_generation_id,
        target_generation_fence=grant.target_generation_fence,
        component_stage_id=stage.component_stage_id,
        source_revision=stage.source_revision,
        source_manifest_hash=stage.source_manifest_hash,
        source_head_fence=stage.source_head_fence,
        operation_id="release_1",
    )
    middle = repository.create_staging([], generation_id="generation_middle")
    repository.release(middle.generation_id)
    final = repository.create_staging([], generation_id="generation_final")
    repository.release(final.generation_id)
    with engine.begin() as connection:
        repository.record_published_chunks(
            _request(generation=grant.target_generation_id),
            [_chunk("graph_gc_chunk", generation=grant.target_generation_id)],
            connection=connection,
        )
    source.candidate_generation_id = grant.target_generation_id
    accepted = coordinator.request_index_generation_gc(
        candidate_generation_id=grant.target_generation_id,
        reconciliation_run_id="reconcile_1",
        operation_id="gc_1",
        caller_principal="retention-ops",
    )
    reader_fence_states: list[tuple[str, tuple[str, ...]]] = []
    stop_graph_readers = repository.stop_graph_readers

    def record_reader_fence(candidate_generation_id, *, connection):
        stop_graph_readers(candidate_generation_id, connection=connection)
        graph = repository.get_generation(candidate_generation_id, connection=connection).manifest[
            "components"
        ]["public_graph"]
        reader_fence_states.append((str(graph["state"]), tuple(graph["graph_resource_ids"])))

    repository.stop_graph_readers = record_reader_fence

    completed = coordinator.complete_index_generation_gc(
        candidate_generation_id=grant.target_generation_id,
        operation_id="gc_1",
        caller_principal="retention-ops",
    )

    assert accepted.state == "accepted"
    assert completed.state == "already_purged"
    assert reader_fence_states == [("disabled", ("graph_resource_1",))]
    assert source.events == [("disabled", "", (), 1)]
    assert manager.active_generation_id == final.generation_id
    with engine.connect() as connection:
        assert (
            connection.execute(
                select(index_chunks_table.c.id).where(
                    index_chunks_table.c.generation_id == grant.target_generation_id
                )
            ).all()
            == []
        )


def test_durable_repository_gc_cannot_bypass_ready_graph_component_cleanup() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    documents_metadata.create_all(engine)
    indexing_metadata.create_all(engine)
    source = PublicGraphSourceService(
        engine,
        trusted_consumers={"indexing": {"indexer_1"}},
        outbox_port=_Outbox(),
    )
    snapshot = source.record_source_change(
        space_id="public", document_id="document_1", change_type="publish", publications=[]
    )
    repository = SqlAlchemyIndexingRepository(engine)
    manager = SqlAlchemyGenerationManager(repository)
    coordinator = GraphComponentCoordinator(manager, source, consumer_id="indexer_1")
    grant = coordinator.reserve_graph_component_stage(
        graph_build_id="graph_1",
        expected_source_revision=snapshot.source_revision,
        expected_active_generation_id=manager.active_generation_id,
        operation_id="stage_1",
        component_input=_graph_input(snapshot, operation_id="stage_1"),
    )
    stage = coordinator.stage_public_graph_component(
        grant,
        graph_resource_manifest_hash="graph_manifest_1",
        graph_resource_ids=["graph_resource_1"],
        build_receipt_hash="build_receipt_1",
    )
    coordinator.release_graph_component(
        target_generation_id=grant.target_generation_id,
        target_generation_fence=grant.target_generation_fence,
        component_stage_id=stage.component_stage_id,
        source_revision=stage.source_revision,
        source_manifest_hash=stage.source_manifest_hash,
        source_head_fence=stage.source_head_fence,
        operation_id="release_1",
    )
    middle = repository.create_staging([], generation_id="generation_middle")
    repository.release(middle.generation_id)
    final = repository.create_staging([], generation_id="generation_final")
    repository.release(final.generation_id)
    repository.request_index_generation_gc(
        grant.target_generation_id, reconciliation_run_id="reconcile_1", operation_id="gc_1"
    )

    result = repository.complete_generation_gc(grant.target_generation_id, operation_id="gc_1")

    assert result.state == "blocked"
    assert result.blocking_reasons == ("graph_component_cleanup_required",)
    assert repository.get_generation(grant.target_generation_id).status == "retired"
    assert manager.active_generation_id == final.generation_id


def test_durable_graph_gc_cleanup_failure_is_retryable_and_does_not_purge() -> None:
    class FailingDiscardSource(PublicGraphSourceService):
        def acknowledge_consumption(self, **kwargs):
            if kwargs["purpose"] == "discard":
                raise PlatformError("discard_transport_failed", "component cleanup failed", {}, 503)
            return super().acknowledge_consumption(**kwargs)

    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    documents_metadata.create_all(engine)
    indexing_metadata.create_all(engine)
    source = FailingDiscardSource(
        engine,
        trusted_consumers={"indexing": {"indexer_1"}},
        outbox_port=_Outbox(),
    )
    snapshot = source.record_source_change(
        space_id="public", document_id="document_1", change_type="publish", publications=[]
    )
    repository = SqlAlchemyIndexingRepository(engine)
    manager = SqlAlchemyGenerationManager(repository)
    coordinator = GraphComponentCoordinator(manager, source, consumer_id="indexer_1")
    grant = coordinator.reserve_graph_component_stage(
        graph_build_id="graph_1",
        expected_source_revision=snapshot.source_revision,
        expected_active_generation_id=manager.active_generation_id,
        operation_id="stage_1",
        component_input=_graph_input(snapshot, operation_id="stage_1"),
    )
    stage = coordinator.stage_public_graph_component(
        grant,
        graph_resource_manifest_hash="graph_manifest_1",
        graph_resource_ids=["graph_resource_1"],
        build_receipt_hash="build_receipt_1",
    )
    coordinator.release_graph_component(
        target_generation_id=grant.target_generation_id,
        target_generation_fence=grant.target_generation_fence,
        component_stage_id=stage.component_stage_id,
        source_revision=stage.source_revision,
        source_manifest_hash=stage.source_manifest_hash,
        source_head_fence=stage.source_head_fence,
        operation_id="release_1",
    )
    middle = repository.create_staging([], generation_id="generation_middle")
    repository.release(middle.generation_id)
    final = repository.create_staging([], generation_id="generation_final")
    repository.release(final.generation_id)
    with engine.begin() as connection:
        repository.record_published_chunks(
            _request(generation=grant.target_generation_id),
            [_chunk("graph_gc_chunk", generation=grant.target_generation_id)],
            connection=connection,
        )
    coordinator.request_index_generation_gc(
        candidate_generation_id=grant.target_generation_id,
        reconciliation_run_id="reconcile_1",
        operation_id="gc_1",
        caller_principal="retention-ops",
    )

    result = coordinator.complete_index_generation_gc(
        candidate_generation_id=grant.target_generation_id,
        operation_id="gc_1",
        caller_principal="retention-ops",
    )

    assert result.state == "blocked"
    assert result.blocking_reasons == ("discard_transport_failed",)
    assert result.retryable is True
    candidate = manager.get_generation(grant.target_generation_id)
    assert candidate.status == "retired"
    assert candidate.manifest["components"]["public_graph"]["state"] == "ready"
    assert candidate.manifest["components"]["public_graph"]["graph_resource_ids"] == [
        "graph_resource_1"
    ]
    assert manager.active_generation_id == final.generation_id
    assert (
        coordinator.request_index_generation_gc(
            candidate_generation_id=grant.target_generation_id,
            reconciliation_run_id="reconcile_1",
            operation_id="gc_1",
            caller_principal="retention-ops",
        ).state
        == "accepted"
    )
