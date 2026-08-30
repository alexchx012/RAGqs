"""Value objects and rendering projection for document previews."""

from __future__ import annotations

import csv
import io
import json
from collections.abc import Mapping
from dataclasses import dataclass, field
from datetime import date, datetime, time
from typing import Any, Protocol
from zipfile import BadZipFile

from openpyxl import load_workbook  # type: ignore[import-untyped]

from app.platform.errors import PlatformError


@dataclass(frozen=True, slots=True)
class PreviewMetadata:
    has_text_layer: bool
    tree_indexed: bool
    page_count: int | None
    sheets: tuple[Mapping[str, Any], ...] | None

    def to_mapping(self) -> dict[str, Any]:
        return {
            "has_text_layer": self.has_text_layer,
            "tree_indexed": self.tree_indexed,
            "page_count": self.page_count,
            "sheets": [dict(sheet) for sheet in self.sheets] if self.sheets is not None else None,
        }


@dataclass(frozen=True, slots=True)
class PreviewHit:
    index: int
    summary: str
    locator: Mapping[str, Any]
    snippet: str | None = None

    def to_mapping(self) -> dict[str, Any]:
        value: dict[str, Any] = {
            "index": self.index,
            "summary": self.summary,
            "locator": dict(self.locator),
        }
        if self.snippet:
            value["snippet"] = self.snippet
        return value


@dataclass(frozen=True, slots=True)
class PreviewContent:
    body: bytes
    media_type: str
    status_code: int = 200
    headers: Mapping[str, str] = field(default_factory=dict)


class PreviewRendererPort(Protocol):
    def metadata(self, *, processing_summary: Mapping[str, Any]) -> PreviewMetadata: ...

    def render(
        self,
        *,
        version: Mapping[str, Any],
        content: bytes,
        metadata: PreviewMetadata,
        sheet: str | None,
    ) -> PreviewContent: ...


def preview_media_kind(media_kind: object) -> str:
    value = str(media_kind or "").split(";", 1)[0].strip().casefold()
    if value in {"application/pdf", "pdf"}:
        return "pdf"
    if value.startswith("image/") or value == "image":
        return "image"
    if value in {
        "application/msword",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "docx",
        "word",
    }:
        return "word"
    if value in {"text/csv", "csv"} or value.endswith("csv"):
        return "csv"
    if value in {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
        "xlsx",
        "xls",
        "excel",
    }:
        return "excel"
    if value in {"text/markdown", "md", "markdown"}:
        return "md"
    if value in {"text/plain", "txt", "text"}:
        return "txt"
    if value in {"text/x-python", "python", "code", "text/javascript", "javascript"}:
        return "code"
    if value in {
        "application/json",
        "json",
        "application/yaml",
        "text/yaml",
        "yaml",
        "application/toml",
        "toml",
        "application/xml",
        "text/xml",
        "xml",
        "data",
    }:
        return "data"
    return "txt"


def is_raw_preview_content(media_type: str) -> bool:
    return preview_media_kind(media_type) in {"pdf", "image"}


def is_pdf_preview_content(media_type: str) -> bool:
    return preview_media_kind(media_type) == "pdf"


def parse_single_byte_range(value: str, size: int) -> tuple[int, int] | None:
    """Return a satisfiable single byte range or ``None`` for an invalid request."""

    if size < 1:
        return None
    header = value.strip()
    if not header.startswith("bytes="):
        return None
    specifier = header.removeprefix("bytes=")
    if not specifier or "," in specifier or specifier.count("-") != 1:
        return None
    start_value, end_value = specifier.split("-", 1)
    if not start_value:
        if not end_value.isdecimal():
            return None
        suffix_length = int(end_value)
        if suffix_length < 1:
            return None
        return max(0, size - suffix_length), size - 1
    if not start_value.isdecimal():
        return None
    start = int(start_value)
    if start >= size:
        return None
    if not end_value:
        return start, size - 1
    if not end_value.isdecimal():
        return None
    end = min(int(end_value), size - 1)
    if end < start:
        return None
    return start, end


class ProcessingReceiptPreviewRenderer:
    """Projects only persisted receipt facts needed by the browser preview."""

    def metadata(self, *, processing_summary: Mapping[str, Any]) -> PreviewMetadata:
        page_count = processing_summary.get("page_count")
        if isinstance(page_count, bool) or not isinstance(page_count, int) or page_count < 1:
            page_count = None
        tree = processing_summary.get("tree")
        tree_indexed = bool(tree.get("tree_indexed")) if isinstance(tree, Mapping) else False
        sheet_manifest = processing_summary.get("sheet_manifest")
        row_groups = processing_summary.get("row_groups")
        row_counts: dict[str, int] = {}
        if isinstance(row_groups, list):
            for group in row_groups:
                if not isinstance(group, Mapping):
                    continue
                sheet = group.get("sheet")
                end = group.get("end")
                if (
                    isinstance(sheet, str)
                    and sheet
                    and isinstance(end, int)
                    and not isinstance(end, bool)
                ):
                    row_counts[sheet] = max(row_counts.get(sheet, 0), end)
        sheets: list[dict[str, Any]] = []
        if isinstance(sheet_manifest, list):
            for item in sheet_manifest:
                if not isinstance(item, Mapping):
                    continue
                sheet = item.get("sheet")
                if isinstance(sheet, str) and sheet:
                    manifest_row_count = item.get("row_count")
                    row_count = (
                        manifest_row_count
                        if isinstance(manifest_row_count, int)
                        and not isinstance(manifest_row_count, bool)
                        and manifest_row_count >= 0
                        else row_counts.get(sheet, 0)
                    )
                    sheets.append({"name": sheet, "row_count": row_count})
        return PreviewMetadata(
            has_text_layer=bool(processing_summary.get("has_text_layer")),
            tree_indexed=tree_indexed,
            page_count=page_count,
            sheets=tuple(sheets) if sheets else None,
        )

    def render(
        self,
        *,
        version: Mapping[str, Any],
        content: bytes,
        metadata: PreviewMetadata,
        sheet: str | None,
    ) -> PreviewContent:
        kind = preview_media_kind(version.get("media_kind"))
        if kind == "pdf":
            return PreviewContent(
                body=content,
                media_type="application/pdf",
                headers={"Accept-Ranges": "bytes"},
            )
        if kind == "image":
            media_type = str(version.get("media_kind") or "application/octet-stream")
            return PreviewContent(body=content, media_type=media_type)
        if kind == "word":
            sections = self._word_sections(version) or []
            if metadata.tree_indexed:
                return self._json_content({"sections": sections})
            return PreviewContent(
                body=self._word_text(sections).encode("utf-8"), media_type="text/plain"
            )
        if kind == "csv":
            return self._csv_content(content, sheet)
        if kind == "excel":
            return self._xlsx_content(content, sheet)
        return PreviewContent(
            body=content.decode("utf-8", errors="replace").encode("utf-8"), media_type="text/plain"
        )

    @staticmethod
    def _word_sections(version: Mapping[str, Any]) -> list[dict[str, list[str]]] | None:
        summary = version.get("processing_summary")
        if not isinstance(summary, Mapping):
            return None
        tree = summary.get("tree")
        if not isinstance(tree, Mapping):
            return None
        source_sections = tree.get("sections")
        if not isinstance(source_sections, list):
            return None
        sections: list[dict[str, list[str]]] = []
        for source in source_sections:
            if not isinstance(source, Mapping):
                continue
            path = source.get("path")
            paragraphs = source.get("paragraphs")
            if (
                not isinstance(path, list)
                or not all(isinstance(part, str) for part in path)
                or not isinstance(paragraphs, list)
                or not all(isinstance(paragraph, str) for paragraph in paragraphs)
            ):
                continue
            sections.append({"path": list(path), "paragraphs": list(paragraphs)})
        return sections

    @staticmethod
    def _word_text(sections: list[dict[str, list[str]]]) -> str:
        blocks = [
            "\n".join([*section["path"], *section["paragraphs"]]).strip() for section in sections
        ]
        return "\n\n".join(block for block in blocks if block)

    @classmethod
    def _csv_content(cls, content: bytes, sheet: str | None) -> PreviewContent:
        if sheet not in {None, "CSV"}:
            raise PlatformError("sheet_not_found", "Sheet was not found", {"sheet": sheet}, 404)
        rows = [list(row) for row in csv.reader(io.StringIO(content.decode("utf-8-sig")))]
        return cls._json_content({"sheet": "CSV", "row_count": len(rows), "rows": rows})

    @classmethod
    def _xlsx_content(cls, content: bytes, sheet: str | None) -> PreviewContent:
        try:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except (BadZipFile, KeyError, OSError, ValueError) as exc:
            raise PlatformError(
                "document_content_unavailable", "Document content is unavailable", {}, 410
            ) from exc
        try:
            selected_sheet = sheet or (workbook.sheetnames[0] if workbook.sheetnames else None)
            if selected_sheet is None or selected_sheet not in workbook.sheetnames:
                raise PlatformError(
                    "sheet_not_found", "Sheet was not found", {"sheet": selected_sheet}, 404
                )
            worksheet = workbook[selected_sheet]
            rows = [
                [cls._json_cell(value) for value in row]
                for row in worksheet.iter_rows(values_only=True)
            ]
        finally:
            workbook.close()
        return cls._json_content({"sheet": selected_sheet, "row_count": len(rows), "rows": rows})

    @staticmethod
    def _json_cell(value: Any) -> str | int | float | bool | None:
        if value is None or isinstance(value, (str, int, float, bool)):
            return value
        if isinstance(value, (date, datetime, time)):
            return value.isoformat()
        return str(value)

    @staticmethod
    def _json_content(value: Mapping[str, Any]) -> PreviewContent:
        return PreviewContent(
            body=json.dumps(value, ensure_ascii=False, separators=(",", ":")).encode("utf-8"),
            media_type="application/json",
        )
