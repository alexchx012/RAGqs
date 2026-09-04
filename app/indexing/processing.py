from __future__ import annotations

import ast
import csv
import hashlib
import io
import json
import posixpath
import re
import tomllib
import zipfile
from bisect import bisect_left, bisect_right
from collections.abc import Callable, Mapping, Sequence
from dataclasses import dataclass, replace
from datetime import UTC, datetime, timedelta
from difflib import SequenceMatcher
from typing import Any, Protocol
from xml.etree import ElementTree

from app.documents.indexing import IndexProcessingReceipt, IndexStagingRequest
from app.platform.errors import PlatformError
from app.usage.ledger import LocalMeasurement, OwnershipSnapshot, ProviderMeasurement

from .contextual import (
    CONTEXTUAL_PREFIX_TOKEN_LIMIT,
    CONTEXTUAL_PROMPT_SCHEMA_VERSION,
    CONTEXTUAL_PROVIDER_ATTEMPTS,
    CONTEXTUAL_TOKENIZATION_VERSION,
    ContextualDocument,
    ContextualRetrievalProvider,
    ContextualRetrievalService,
    ContextualUsageFact,
    approximate_token_count,
    contextual_target,
)
from .image_vlm import (
    image_dimensions,
    is_decorative,
    normalize_description,
)
from .models import IndexChunk
from .prefix_cache import PrefixCacheManager
from .profiles import document_profile_for_media_kind

# OCR pages below this confidence are flagged low-confidence in the processing
# receipt; the value is a stable default, not per-document configuration.
OCR_LOW_CONFIDENCE_THRESHOLD = 0.9
CONTINUATION_HEADER_SIMILARITY = 0.7
MINERU_PROVIDER_ATTEMPTS = 5
# 结构化分流判据（A61）：schema 特征组合（单独的顶层 "type" 不再判 schema）；
# 深嵌套阈值（容器嵌套层数）——达到阈值的无 schema 键对象改走代码/配置路径，
# 不再做键值对扁平化。
_STRUCTURED_SCHEMA_KEYS: tuple[str, ...] = (
    "$schema",
    "schema",
    "properties",
    "required",
    "definitions",
    "$defs",
)
STRUCTURED_NESTING_DEPTH_THRESHOLD = 3


class CompressionPort(Protocol):
    def compress(self, text: str, *, context: Mapping[str, Any]) -> str: ...


class UsageSubmissionPort(Protocol):
    def submit_local_usage(
        self,
        *,
        execution_kind: str,
        execution_id: str,
        stage: str,
        resource_kind: str,
        measurement: LocalMeasurement,
        ownership: OwnershipSnapshot,
        result: str,
        started_at_utc: datetime,
        replay_generation: int = 0,
    ) -> str | None: ...


class IdentityCompression:
    def compress(self, text: str, *, context: Mapping[str, Any]) -> str:
        del context
        return text.strip()


@dataclass(frozen=True, slots=True)
class OCRSamplePlan:
    page_count: int
    pages: tuple[int, ...]

    @classmethod
    def for_page_count(cls, page_count: int) -> OCRSamplePlan:
        pages: tuple[int, ...]
        if page_count < 1:
            return cls(0, ())
        if page_count <= 10:
            pages = (1, page_count) if page_count > 1 else (1,)
        elif page_count <= 50:
            pages = (1, 2, max(1, page_count // 2), max(1, page_count // 2 + 1), page_count)
        elif page_count <= 200:
            pages = (
                tuple(range(1, 6))
                + tuple(max(1, page_count // 2 - 1 + offset) for offset in range(3))
                + tuple(range(max(1, page_count - 1), page_count + 1))
            )
        else:
            pages = (
                tuple(range(1, 9))
                + tuple(max(1, page_count // 2 - 1 + offset) for offset in range(3))
                + tuple(range(max(1, page_count - 2), page_count + 1))
            )
        return cls(page_count, tuple(dict.fromkeys(pages)))


@dataclass(frozen=True, slots=True)
class ProcessingOutput:
    chunks: tuple[IndexChunk, ...]
    receipt: IndexProcessingReceipt


# ---------------------------------------------------------------------------
# 交叉引用反向链接解析（07-4.6.9）
# ---------------------------------------------------------------------------

_CROSS_REF_PATTERNS: tuple[re.Pattern[str], ...] = tuple(
    re.compile(pattern, re.IGNORECASE)
    for pattern in (
        r"(?:如|见|参见|参考)\s*图\s*([0-9]+(?:[.:：-][0-9]+)*)",
        r"(?:如|见|参见|参考)\s*表\s*([0-9]+(?:[.:：-][0-9]+)*)",
        r"图\s*([0-9]+(?:[.:：-][0-9]+)*)\s*(?:所示|展示|显示)",
        r"figure\s+([0-9]+(?:[.:：-][0-9]+)*)",
        r"table\s+([0-9]+(?:[.:：-][0-9]+)*)",
        r"(?:如|见|参见)\s*([0-9]+(?:\.[0-9]+)+)\s*(?:节|章)",
        r"section\s+([0-9]+(?:\.[0-9]+)+)",
    )
)


def parse_reverse_links(text: str, *, limit: int = 32) -> tuple[dict[str, str], ...]:
    """解析"如图 X/见图 X/figure N/表 X/§X.Y"类交叉引用。

    返回去重后的目标列表：``{"kind": "figure"|"table"|"section", "ref": "3"}``；
    超出 limit 截断（防解析风暴）。结果写入 chunk ``reverse_links`` 元数据通道，
    供图/表/章节 chunk 建立反向链接（谁引用了我）。
    """

    seen: dict[tuple[str, str], dict[str, str]] = {}
    for pattern in _CROSS_REF_PATTERNS:
        for match in pattern.finditer(text):
            ref = match.group(1).rstrip(".。")
            if not ref:
                continue
            source = match.group(0).lower()
            if "table" in source or "表" in source:
                kind = "table"
            elif "section" in source or "节" in source or "章" in source:
                kind = "section"
            else:
                kind = "figure"
            key = (kind, ref)
            if key not in seen:
                seen[key] = {"kind": kind, "ref": ref}
            if len(seen) >= limit:
                return tuple(seen.values())
    return tuple(seen.values())


def _machine_low_confidence_fact(confidence: float, page: int | None) -> dict[str, Any]:
    """Outbox-closed OCR fact shape: only confidence, page and region."""

    return {"confidence": confidence, "page": int(page or 1), "region": []}


def _group_chunks_by_tree_parent(
    chunks: Sequence[IndexChunk],
) -> list[tuple[str, list[IndexChunk]]]:
    grouped: dict[str, list[IndexChunk]] = {}
    for chunk in chunks:
        if not contextual_target(chunk):
            continue
        path = str(chunk.metadata.get("section_path") or "").strip()
        grouped.setdefault(path, []).append(chunk)
    return list(grouped.items())


def _sha256(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def _text(value: bytes | str) -> str:
    if isinstance(value, bytes):
        return value.decode("utf-8")
    return value


def _heading(line: str) -> bool:
    stripped = line.strip()
    return bool(
        re.match(r"^(?:#{1,6}\s+|(?:\d+\.){1,4}\s+)[^\s].*$", stripped)
        or (stripped and len(stripped) <= 96 and stripped == stripped.title())
        or (
            stripped
            and len(stripped) <= 96
            and stripped.isupper()
            and any(character.isalpha() for character in stripped)
        )
    )


def _sections(text: str) -> list[tuple[str, str]]:
    sections: list[tuple[str, str]] = []
    path: list[str] = []
    body: list[str] = []
    saw_heading = False
    placeholder_number = 0

    def flush() -> None:
        nonlocal body, placeholder_number
        value = "\n".join(body).strip()
        if not value:
            return
        section_path = " / ".join(path)
        if not section_path and saw_heading:
            placeholder_number += 1
            section_path = f"Unsectioned {placeholder_number}"
        sections.append((section_path, value))
        body = []

    for line in text.splitlines():
        if _heading(line):
            flush()
            title = re.sub(r"^\s*(?:#+\s*|(?:\d+\.){1,4}\s+)", "", line).strip()
            level = len(line) - len(line.lstrip("#")) if line.lstrip().startswith("#") else 1
            path = path[: max(0, level - 1)] + [title]
            saw_heading = True
        elif line.strip():
            # B9-locator：段落边界哨兵——空行在原始文本中分隔段落，section
            # 化时以显式空行保留（下游按空行切块/编段落号）；连续空行折叠。
            if body and body[-1] != "":
                body.append("")
            body.append(line.rstrip())
    flush()
    return sections


def _category_column(rows: Sequence[Sequence[str]], width: int) -> int | None:
    for column in range(width):
        values = [row[column].strip() for row in rows if column < len(row) and row[column].strip()]
        if values and len(set(values)) < len(values):
            return column
    return None


def _split_rows(
    rows: Sequence[Sequence[str]], width: int, *, category_column: int | None = None
) -> list[tuple[int, int]]:
    if not rows:
        return []
    size = 30 if width < 20 else 20
    category = _category_column(rows, width) if category_column is None else category_column
    groups: list[tuple[int, int]] = []
    start = 0
    for index in range(1, len(rows)):
        previous, current = rows[index - 1], rows[index]
        changed = False
        if width and category is not None and previous and current:
            previous_value = previous[category].strip() if category < len(previous) else ""
            current_value = current[category].strip() if category < len(current) else ""
            changed = bool(previous_value) and previous_value != current_value
        if index - start >= size or changed:
            groups.append((start, index))
            start = index
    groups.append((start, len(rows)))
    return groups


def _column_name(number: int) -> str:
    result = ""
    while number:
        number, remainder = divmod(number - 1, 26)
        result = chr(65 + remainder) + result
    return result or "A"


def _split_text(value: str, *, maximum: int) -> list[str]:
    return [value[start : start + maximum] for start in range(0, len(value), maximum)]


def _split_lines_preserving_rows(lines: Sequence[str], *, maximum: int) -> list[list[str]]:
    """按整行分组切段（A60）：单行超过上限时整行自成一个块，不切断行内容。"""

    parts: list[list[str]] = []
    current: list[str] = []
    used = 0
    for line in lines:
        step = len(line) + (1 if current else 0)
        if current and used + step > maximum:
            parts.append(current)
            current = [line]
            used = len(line)
        else:
            current.append(line)
            used += step
    if current:
        parts.append(current)
    return parts


def _split_blocks_preserving_tables(value: str, *, maximum: int) -> list[str]:
    """Split text by blank-line blocks; markdown tables stay atomic."""

    blocks = [block for block in re.split(r"\n\s*\n", value) if block.strip()]
    result: list[str] = []
    buffer = ""
    for block in blocks:
        lines = block.splitlines()
        table_block = len(lines) >= 2 and all(line.lstrip().startswith("|") for line in lines)
        if table_block:
            if buffer.strip():
                result.extend(_split_text(buffer, maximum=maximum))
                buffer = ""
            result.append("\n".join(lines))
            continue
        candidate = f"{buffer}\n\n{block}" if buffer else block
        if len(candidate) > maximum and buffer:
            result.extend(_split_text(buffer, maximum=maximum))
            buffer = block
        else:
            buffer = candidate
    if buffer.strip():
        result.extend(_split_text(buffer, maximum=maximum))
    return result or [""]


def _markdown_table(headers: Sequence[str], rows: Sequence[Sequence[str]]) -> str:
    escaped_headers = [str(header).replace("|", "\\|") for header in headers]
    lines = [
        "| " + " | ".join(escaped_headers) + " |",
        "| " + " | ".join("---" for _ in escaped_headers) + " |",
    ]
    for row in rows:
        values = [
            str(row[index] if index < len(row) else "").replace("|", "\\|")
            for index in range(len(escaped_headers))
        ]
        lines.append("| " + " | ".join(values) + " |")
    return "\n".join(lines)


def _key_value_rows(headers: Sequence[str], rows: Sequence[Sequence[str]]) -> str:
    return "\n".join(
        " | ".join(
            f"{header}: {row[index] if index < len(row) else ''}"
            for index, header in enumerate(headers)
        )
        for row in rows
    )


def _header_similarity(left: Sequence[str], right: Sequence[str]) -> float:
    return SequenceMatcher(
        None,
        "\x00".join(str(item).strip().casefold() for item in left),
        "\x00".join(str(item).strip().casefold() for item in right),
    ).ratio()


def _tables_continuation(left: Mapping[str, Any], right: Mapping[str, Any]) -> bool:
    """Continuation check: column count + header similarity + aspect ratio."""

    left_headers = list(left.get("headers") or ())
    right_headers = list(right.get("headers") or ())
    if not left_headers or len(left_headers) != len(right_headers):
        return False
    if _header_similarity(left_headers, right_headers) < CONTINUATION_HEADER_SIMILARITY:
        return False
    left_rows = list(left.get("rows") or ())
    right_rows = list(right.get("rows") or ())
    left_ratio = len(left_rows) / max(len(left_headers), 1)
    right_ratio = len(right_rows) / max(len(right_headers), 1)
    return abs(left_ratio - right_ratio) <= max(left_ratio, right_ratio, 1.0)


def _strip_header_row(headers: Sequence[str], rows: list[list[str]]) -> list[list[str]]:
    if rows and _header_similarity(headers, rows[0]) >= CONTINUATION_HEADER_SIMILARITY:
        return rows[1:]
    return rows


def merge_continuation_tables(
    tables: Sequence[Mapping[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Concatenate continuation tables before row-group chunking."""

    merged: list[dict[str, Any]] = []
    facts: list[dict[str, Any]] = []
    for table in tables:
        headers = [str(item) for item in (table.get("headers") or ())]
        rows = [[str(cell) for cell in row] for row in (table.get("rows") or ())]
        page = int(table.get("page", 1) or 1)
        previous = merged[-1] if merged else None
        if previous and _tables_continuation(previous, {"headers": headers, "rows": rows}):
            incoming = rows
            if (
                incoming
                and _header_similarity(headers, previous["headers"])
                >= CONTINUATION_HEADER_SIMILARITY
            ):
                incoming = incoming[1:]
                facts.append(
                    {
                        "kind": "continuation_header_deduplicated",
                        "pages": [previous["page_end"], page],
                    }
                )
            previous["rows"].extend(incoming)
            previous["page_end"] = page
            continue
        merged.append(
            {
                "headers": headers,
                "rows": _strip_header_row(headers, rows),
                "page_start": page,
                "page_end": page,
                "title": str(table.get("title", "") or ""),
            }
        )
    return merged, facts


def _flatten_multilevel_header(
    header_rows: Sequence[Sequence[str]],
) -> list[str]:
    """Flatten non-empty parent-child header paths (e.g. 财务数据_营收_Q1)."""

    if len(header_rows) <= 1:
        return [str(item).strip() for item in (header_rows[0] if header_rows else ())]
    width = max(len(row) for row in header_rows)
    columns: list[str] = []
    seen: dict[str, int] = {}
    for column in range(width):
        parts: list[str] = []
        for row in header_rows:
            value = row[column].strip() if column < len(row) else ""
            if value and value not in parts:
                parts.append(value)
        name = "_".join(parts) or f"column_{column + 1}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 1
        columns.append(name)
    return columns


def _xlsx_merged_ranges(content: bytes, *, maximum_cells: int) -> dict[str, tuple[str, ...]]:
    from openpyxl.utils.cell import range_boundaries  # type: ignore[import-untyped]

    main_namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    document_rel_namespace = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    package_rel_namespace = "http://schemas.openxmlformats.org/package/2006/relationships"
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        targets: dict[str, str] = {}
        with archive.open("xl/_rels/workbook.xml.rels") as source:
            for _, relation in ElementTree.iterparse(source, events=("end",)):
                if relation.tag == f"{{{package_rel_namespace}}}Relationship":
                    target = relation.attrib["Target"].lstrip("/")
                    if not target.startswith("xl/"):
                        target = posixpath.join("xl", target)
                    targets[relation.attrib["Id"]] = posixpath.normpath(target)
                relation.clear()
        sheets: list[tuple[str, str]] = []
        with archive.open("xl/workbook.xml") as source:
            for _, sheet in ElementTree.iterparse(source, events=("end",)):
                if sheet.tag == f"{{{main_namespace}}}sheet":
                    relationship_id = sheet.attrib.get(f"{{{document_rel_namespace}}}id")
                    if relationship_id is None:
                        raise ValueError("Excel worksheet relationship is missing")
                    sheets.append((sheet.attrib["name"], relationship_id))
                sheet.clear()
        if not sheets:
            raise ValueError("Excel workbook has no sheets")
        result: dict[str, tuple[str, ...]] = {}
        for sheet_name, relationship_id in sheets:
            target = targets.get(relationship_id)
            if target is None:
                raise ValueError("Excel worksheet relationship is missing")
            ranges: list[str] = []
            with archive.open(target) as source:
                for _, merged in ElementTree.iterparse(source, events=("end",)):
                    if merged.tag != f"{{{main_namespace}}}mergeCell":
                        merged.clear()
                        continue
                    reference = merged.attrib["ref"]
                    min_column, min_row, max_column, max_row = range_boundaries(reference)
                    if (max_column - min_column + 1) * (max_row - min_row + 1) > maximum_cells:
                        raise PlatformError(
                            "table_parse_failed",
                            "Excel merged region exceeds the supported size",
                            {},
                            422,
                        )
                    ranges.append(reference)
                    merged.clear()
            result[sheet_name] = tuple(ranges)
    return result


@dataclass(frozen=True, slots=True)
class _MergedRange:
    min_column: int
    min_row: int
    max_column: int
    max_row: int


class _MergedCellResolver:
    def __init__(self, merged_ranges: Sequence[str]) -> None:
        from openpyxl.utils.cell import range_boundaries  # type: ignore[import-untyped]

        self._active_ranges: list[_MergedRange] = []
        self._active_starts: list[int] = []
        self._starting: dict[int, list[_MergedRange]] = {}
        self._ending: dict[int, list[_MergedRange]] = {}
        self.horizontally_merged_rows: set[int] = set()
        for reference in merged_ranges:
            min_column, min_row, max_column, max_row = range_boundaries(reference)
            merged = _MergedRange(min_column, min_row, max_column, max_row)
            self._starting.setdefault(min_row, []).append(merged)
            self._ending.setdefault(max_row + 1, []).append(merged)
            if min_row == max_row and min_column < max_column:
                self.horizontally_merged_rows.add(min_row)

    def advance(self, row_number: int) -> None:
        for merged in self._ending.pop(row_number, ()):
            index = bisect_left(self._active_starts, merged.min_column)
            while self._active_ranges[index] is not merged:
                index += 1
            del self._active_starts[index]
            del self._active_ranges[index]
        for merged in self._starting.pop(row_number, ()):
            index = bisect_right(self._active_starts, merged.min_column)
            self._active_starts.insert(index, merged.min_column)
            self._active_ranges.insert(index, merged)

    def origin_at(self, column: int) -> tuple[int, int] | None:
        index = bisect_right(self._active_starts, column) - 1
        if index < 0:
            return None
        merged = self._active_ranges[index]
        if column > merged.max_column:
            return None
        return (merged.min_row, merged.min_column)


def _flatten(value: Any, prefix: str = "") -> list[tuple[str, str]]:
    if isinstance(value, Mapping):
        result: list[tuple[str, str]] = []
        for key, child in value.items():
            result.extend(_flatten(child, f"{prefix}.{key}" if prefix else str(key)))
        return result
    if isinstance(value, list):
        return [
            (f"{prefix}[{index}]", json.dumps(item, ensure_ascii=False, sort_keys=True))
            for index, item in enumerate(value)
        ]
    return [(prefix, str(value))]


def _nesting_depth(value: Any) -> int:
    """Mapping nesting depth of a parsed payload (scalar = 0).

    列表不额外计层——它承载最深子项的深度（同名兄弟折叠出的列表、重复元素
    不构成用户可感知的嵌套层级）；纯映射链每层 +1。
    """

    if isinstance(value, Mapping):
        return 1 + max((_nesting_depth(child) for child in value.values()), default=0)
    if isinstance(value, list):
        return max((_nesting_depth(child) for child in value), default=0)
    return 0


def _element_to_value(element: ElementTree.Element) -> Any:
    """XML → nested mapping/list value (A61)：保留层级，同名兄弟折叠为列表。"""

    children = list(element)
    if not children:
        return (element.text or "").strip()
    result: dict[str, Any] = {}
    for child in children:
        value = _element_to_value(child)
        existing = result.get(child.tag)
        if isinstance(existing, list):
            existing.append(value)
        elif existing is not None:
            result[child.tag] = [existing, value]
        else:
            result[child.tag] = value
    return result


def _yaml_scalar(raw: str) -> Any:
    value = raw.strip()
    if len(value) >= 2 and value[0] == value[-1] and value[0] in {"'", '"'}:
        return value[1:-1]
    if value.startswith("[") and value.endswith("]"):
        inner = value[1:-1].strip()
        return [item.strip().strip("'\"") for item in inner.split(",")] if inner else []
    return value


def _yaml_key_split(content: str) -> tuple[str, str] | None:
    """Split ``key: value`` / ``key:`` outside of quotes; None when not a key line."""

    for position, character in enumerate(content):
        if character == ":" and (position + 1 == len(content) or content[position + 1] in " \t"):
            key = content[:position].strip().strip("'\"")
            if not key:
                return None
            return key, content[position + 1 :].strip()
    return None


def _parse_yaml_block(lines: list[tuple[int, str]], index: int, indent: int) -> tuple[Any, int]:
    if index >= len(lines):
        return "", index
    if lines[index][1] == "-" or lines[index][1].startswith("- "):
        items: list[Any] = []
        while index < len(lines):
            line_indent, content = lines[index]
            if line_indent != indent or not (content == "-" or content.startswith("- ")):
                break
            rest = content[2:].strip() if content.startswith("- ") else ""
            if not rest:
                if index + 1 < len(lines) and lines[index + 1][0] > indent:
                    value, index = _parse_yaml_block(lines, index + 1, lines[index + 1][0])
                    items.append(value)
                else:
                    items.append("")
                index += 1
            elif _yaml_key_split(rest) is not None:
                # "- key: value"：按 key 的真实列改写为映射行后按映射块解析。
                key_offset = len(content[2:]) - len(content[2:].lstrip(" "))
                key_indent = indent + 2 + key_offset
                spliced = list(lines)
                spliced[index] = (key_indent, rest)
                value, index = _parse_yaml_block(spliced, index, key_indent)
                items.append(value)
            else:
                items.append(_yaml_scalar(rest))
                index += 1
        return items, index
    mapping: dict[str, Any] = {}
    while index < len(lines):
        line_indent, content = lines[index]
        if line_indent != indent:
            break
        split = _yaml_key_split(content)
        if split is None:
            raise ValueError(f"yaml mapping line is invalid: {content}")
        key, rest = split
        if rest:
            mapping[key] = _yaml_scalar(rest)
            index += 1
        elif index + 1 < len(lines) and lines[index + 1][0] > indent:
            value, index = _parse_yaml_block(lines, index + 1, lines[index + 1][0])
            mapping[key] = value
        else:
            mapping[key] = ""
            index += 1
    return mapping, index


def _parse_yaml_subset(text: str) -> Any:
    """无第三方依赖的 YAML 子集解析（A61）：嵌套映射、块列表与流式标量列表。

    覆盖配置文件的常用形态（缩进映射、``- `` 列表、``- key: value`` 列表项、
    引号标量、``[a, b]`` 流式列表、``#`` 注释与 ``---`` 分隔符）；结构不一致时
    抛出 ValueError（由结构化入口统一映射为 ``structured_parse_failed``）。
    """

    lines: list[tuple[int, str]] = []
    for raw_line in text.splitlines():
        stripped = raw_line.strip()
        if not stripped or stripped.startswith("#") or stripped in {"---", "..."}:
            continue
        indent = len(raw_line) - len(raw_line.lstrip(" "))
        if raw_line[indent : indent + 1] == "\t":
            raise ValueError("yaml indentation must use spaces")
        lines.append((indent, stripped))
    if not lines:
        return {}
    value, consumed = _parse_yaml_block(lines, 0, lines[0][0])
    if consumed != len(lines):
        raise ValueError("yaml structure is inconsistent")
    return value


class ContentProcessor:
    """Deterministic format routing and derivation boundary.

    Expensive OCR/VLM/CR work is injected through ports; the default adapter is
    intentionally only suitable for development and contract tests.
    """

    def __init__(
        self,
        *,
        compressor: CompressionPort | None = None,
        mineru: Callable[[bytes], Mapping[str, Any]] | None = None,
        image_describer: Callable[[bytes, Mapping[str, Any]], Any] | None = None,
        image_ocr: Callable[[bytes, Mapping[str, Any]], str] | None = None,
        text_chunk_max_chars: int = 8_000,
        xlsx_merged_cells_max: int = 10_000,
        ocr_confidence_threshold: float = 0.9,
        usage_submission: UsageSubmissionPort | None = None,
        contextual_provider: ContextualRetrievalProvider | None = None,
        contextual_cache: PrefixCacheManager | None = None,
        contextual_concurrency: int = 4,
        contextual_max_attempts: int = CONTEXTUAL_PROVIDER_ATTEMPTS,
        contextual_prefix_token_limit: int = CONTEXTUAL_PREFIX_TOKEN_LIMIT,
        contextual_token_counter: Callable[[str], int] | None = None,
        model_version: str = "none",
        prompt_version: str = "none",
    ) -> None:
        self._compressor = compressor or IdentityCompression()
        self._mineru = mineru
        self._image_describer = image_describer
        self._image_ocr = image_ocr
        self._text_chunk_max_chars = text_chunk_max_chars
        self._xlsx_merged_cells_max = xlsx_merged_cells_max
        self._ocr_confidence_threshold = float(ocr_confidence_threshold)
        self._usage_submission = usage_submission
        self._contextual_provider = contextual_provider
        self._contextual_cache = contextual_cache
        self._contextual_concurrency = contextual_concurrency
        self._contextual_max_attempts = contextual_max_attempts
        self._contextual_prefix_token_limit = contextual_prefix_token_limit
        self._contextual_token_counter = contextual_token_counter
        self._model_version = model_version
        self._prompt_version = prompt_version

    def processing_identity(self) -> dict[str, str]:
        """Effective model/prompt identity for replay snapshot freezing (§2.3).

        The manual replay transaction freezes these values into the job's
        processing config snapshot, so every attempt of the same replay
        generation stamps the identical model/prompt identity into its
        receipt instead of re-resolving whatever is live at execution time.
        """
        identity = {"model_version": self._model_version, "prompt_version": self._prompt_version}
        if self._contextual_provider is not None:
            identity["cr_model"] = str(self._contextual_provider.model)
        return identity

    def process(
        self,
        request: IndexStagingRequest,
        content: bytes | str,
        *,
        media_kind: str,
        content_manifest_id: str,
        content_manifest_hash: str,
        processing_config_version: str | None = None,
        model_version: str | None = None,
        prompt_version: str | None = None,
        page_count: int = 0,
        has_text_layer: bool | None = None,
        image_context: Mapping[str, Any] | None = None,
        image_ocr_text: str | None = None,
        decorative_image: bool = False,
    ) -> ProcessingOutput:
        raw = content if isinstance(content, bytes) else content.encode("utf-8")
        model_version = self._model_version if model_version is None else model_version
        prompt_version = self._prompt_version if prompt_version is None else prompt_version
        kind = media_kind.strip().casefold()
        document_profile = document_profile_for_media_kind(media_kind)
        profile = processing_config_version or document_profile.config_version
        route_adapter = "text"
        local_usage_facts: list[dict[str, Any]] = []
        document_text = ""
        mineru_probe_failed = False
        if not content_manifest_id or not content_manifest_hash:
            raise PlatformError(
                "validation_error", "content manifest identity is required", {}, 422
            )
        if request.input_manifest_hash != content_manifest_hash:
            raise PlatformError(
                "processing_receipt_conflict",
                "content manifest does not match the staging request",
                {},
                409,
            )
        if kind in {"text/plain", "text/markdown", "txt", "md", "markdown"}:
            document_text = _text(content)
            chunks, summary, degradations = self._text_chunks(
                request, document_text, media_kind, content_manifest_hash
            )
        elif kind in {"text/csv", "csv"} or kind.endswith("csv"):
            document_text = _text(content)
            chunks, summary, degradations = self._csv_chunks(
                request, document_text, content_manifest_hash
            )
        elif kind in {
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
            "xlsx",
            "excel",
        }:
            route_adapter = "structured-table"
            chunks, summary, degradations = self._xlsx_chunks(request, raw, content_manifest_hash)
        elif kind in {
            "application/json",
            "json",
            "application/yaml",
            "text/yaml",
            "yaml",
            "application/toml",
            "toml",
            "application/xml",
            "text/xml",
            "xml",
        }:
            document_text = _text(content)
            chunks, summary, degradations = self._structured_chunks(
                request, document_text, kind, content_manifest_hash
            )
        elif kind in {"text/x-python", "python", "code", "text/javascript", "javascript"}:
            document_text = _text(content)
            chunks, summary, degradations = self._code_chunks(
                request, document_text, content_manifest_hash
            )
        elif kind in {
            "application/pdf",
            "pdf",
            "application/msword",
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "docx",
            "word",
            "application/vnd.ms-powerpoint",
            "application/vnd.openxmlformats-officedocument.presentationml.presentation",
            "ppt",
            "pptx",
            "powerpoint",
            "text/html",
            "html",
            "url",
            "web_url",
        }:
            route_adapter = "mineru-pipeline"
            if self._mineru is None:
                raise PlatformError(
                    "processor_unavailable", "MinerU processor is not configured", {}, 503
                )
            parse = getattr(self._mineru, "parse", None)
            try:
                for _attempt in range(1, MINERU_PROVIDER_ATTEMPTS + 1):
                    try:
                        if callable(parse):
                            parsed = parse(
                                raw,
                                media_kind=media_kind,
                                page_count=page_count,
                                source_url=media_kind if kind in {"url", "web_url"} else None,
                            )
                        else:
                            parsed = self._mineru(raw)
                        break
                    except PlatformError as exc:
                        if (
                            exc.code != "mineru_parse_failed"
                            or _attempt == MINERU_PROVIDER_ATTEMPTS
                        ):
                            raise
            except PlatformError as exc:
                if exc.code != "mineru_parse_failed":
                    raise
                mineru_probe_failed = True
                parsed = {
                    "text": raw.decode("latin-1"),
                    "page_count": page_count or 1,
                    "has_text_layer": False,
                    "structure": {"class": "basic", "reason": "ocr_probe_unavailable"},
                    "backend": "pipeline",
                }
            parsed_text = str(parsed.get("text", ""))
            if not parsed_text.strip():
                raise PlatformError("processing_failed", "MinerU returned no text", {}, 422)
            document_text = parsed_text
            chunks, summary, degradations = self._text_chunks(
                request, parsed_text, media_kind, content_manifest_hash
            )
            summary = dict(summary)
            page_count = int(parsed.get("page_count", page_count or 0))
            summary["page_count"] = page_count
            # B4 内嵌图片计量：MinerU figure 块计数进入 summary；bailian 折算
            # 扣页（文档载体 pages+images），internvl 只记用量不扣（消费端
            # _metered_pages 按 embedded_image_provider 区分）。降级路径
            # （mineru_probe_failed）不带 image_count → 不虚计。
            embedded_images = parsed.get("image_count")
            if (
                isinstance(embedded_images, int)
                and not isinstance(embedded_images, bool)
                and embedded_images > 0
            ):
                summary["image_count"] = embedded_images
                provider = getattr(self._image_describer, "provider", None)
                summary["metering"] = {
                    "class": "document",
                    "embedded_image_count": embedded_images,
                    "embedded_image_provider": str(provider) if provider else "none",
                }
            parsed_text_layer = parsed.get("has_text_layer")
            has_text_layer = (
                bool(parsed_text_layer) if parsed_text_layer is not None else bool(has_text_layer)
            )
            ocr = dict(summary.get("ocr", {}))
            confidence = parsed.get("ocr_confidence")
            if confidence is not None:
                ocr["confidence"] = float(confidence)
                ocr["low_confidence"] = float(confidence) < self._ocr_confidence_threshold
                ocr["reason"] = "low_confidence"
                ocr["status"] = "degraded" if ocr["low_confidence"] else "ok"
                ocr["fact"] = _machine_low_confidence_fact(float(confidence), None)
                ocr["threshold"] = self._ocr_confidence_threshold
                ocr["threshold_version"] = f"ocr_threshold:{self._ocr_confidence_threshold}"
            local_confidence = parsed.get("ocr_confidence_by_page", {})
            if isinstance(local_confidence, Mapping):
                sample = OCRSamplePlan.for_page_count(page_count)
                sampled = {
                    str(page): float(
                        local_confidence[str(page)]
                        if str(page) in local_confidence
                        else local_confidence[page]
                    )
                    for page in sample.pages
                    if str(page) in local_confidence or page in local_confidence
                }
                if sampled:
                    lowest = min(sampled.values())
                    worst_page = min(
                        int(page) for page, value in sampled.items() if value == lowest
                    )
                    ocr["local_confidence"] = sampled
                    ocr["low_confidence"] = lowest < self._ocr_confidence_threshold
                    ocr["reason"] = "low_confidence"
                    ocr["status"] = "degraded" if ocr["low_confidence"] else "ok"
                    ocr["fact"] = _machine_low_confidence_fact(lowest, worst_page)
                    ocr["threshold"] = self._ocr_confidence_threshold
                    ocr["threshold_version"] = f"ocr_threshold:{self._ocr_confidence_threshold}"
            summary["ocr"] = ocr
            structure = parsed.get("structure")
            structure_class = (
                str(structure.get("class", "tree")) if isinstance(structure, Mapping) else "tree"
            )
            if isinstance(structure, Mapping) and structure.get("signal_low_confidence"):
                signal_values = [
                    float(value)
                    for value in (structure.get("signal_confidence") or {}).values()
                    if value is not None
                ]
                if signal_values:
                    lowest_signal = min(signal_values)
                    ocr["low_confidence"] = True
                    ocr["reason"] = "structure_low_confidence"
                    ocr["status"] = "degraded"
                    ocr["fact"] = _machine_low_confidence_fact(lowest_signal, None)
                    ocr["threshold"] = self._ocr_confidence_threshold
                    ocr["threshold_version"] = f"ocr_threshold:{self._ocr_confidence_threshold}"
            if structure_class == "basic":
                chunks = [
                    replace(
                        chunk,
                        locator={},
                        metadata={**dict(chunk.metadata), "section_path": ""},
                    )
                    for chunk in chunks
                ]
                summary["tree"] = {
                    "tree_indexed": False,
                    "tree_reason": "no_structure",
                    "structure_class": structure_class,
                }
            elif structure_class == "partial":
                chunks = [
                    replace(
                        chunk,
                        metadata={
                            **dict(chunk.metadata),
                            "section_path": str(chunk.metadata.get("section_path") or "")
                            or "此章未能分类",
                        },
                    )
                    for chunk in chunks
                ]
                tree = dict(summary.get("tree") or {})
                tree["structure_class"] = structure_class
                summary["tree"] = tree
            else:
                tree = dict(summary.get("tree") or {})
                tree["structure_class"] = structure_class
                summary["tree"] = tree
            parsed_usage = parsed.get("usage")
            if isinstance(parsed_usage, Mapping) and parsed_usage.get("kind") == "local_usage":
                local_usage_facts.append(dict(parsed_usage))
            parsed_timings = parsed.get("timings")
            if isinstance(parsed_timings, Mapping):
                summary["mineru_timings"] = {
                    "sample_ms": int(parsed_timings.get("sample_ms", 0) or 0),
                    "full_ms": int(parsed_timings.get("full_ms", 0) or 0),
                }
            parsed_chunks = parsed.get("chunks", ())
            locations = tuple(parsed_chunks) if isinstance(parsed_chunks, Sequence) else ()
            # B9 消费端精化：页重建文本（mineru page_texts）供 snippet 页内定位。
            raw_page_texts = parsed.get("page_texts")
            page_texts = (
                {str(key): str(value) for key, value in raw_page_texts.items()}
                if isinstance(raw_page_texts, Mapping)
                else {}
            )

            def _pdf_span(index: int, chunk: IndexChunk) -> str | None:
                """页内字符偏移：优先 snippet 在页重建文本中的真实定位
                （同页多处重复可消歧）；未命中回退 mineru 的页文本全幅 span。"""

                location = (
                    locations[index]
                    if index < len(locations) and isinstance(locations[index], Mapping)
                    else None
                )
                if location is None or not has_text_layer:
                    return None
                fallback = location.get("span")
                page_number = str(int(location.get("page", 1)))
                page_text = page_texts.get(page_number, "")
                snippet = chunk.snippet or ""
                if page_text and snippet:
                    probe = snippet[:80]
                    found = page_text.find(probe)
                    if found >= 0:
                        return f"{found}:{found + len(snippet)}"
                return str(fallback) if fallback is not None else None

            if kind in {"application/pdf", "pdf"}:
                if structure_class == "basic":
                    pass
                else:
                    chunks = [
                        replace(
                            chunk,
                            locator=(
                                {
                                    "page": int(locations[index].get("page", 1)),
                                    **(
                                        {"span": span_value}
                                        if (span_value := _pdf_span(index, chunk)) is not None
                                        else {}
                                    ),
                                }
                                if index < len(locations) and isinstance(locations[index], Mapping)
                                else {"page": min(index + 1, max(page_count, 1))}
                            ),
                            snippet=chunk.snippet if has_text_layer else None,
                        )
                        for index, chunk in enumerate(chunks)
                    ]
            else:
                chunks = [
                    replace(
                        chunk,
                        snippet=None,
                    )
                    for chunk in chunks
                ]
                tree = dict(summary.get("tree") or {})
                sections: list[dict[str, list[str]]] = []
                for chunk in chunks:
                    section_path = str(chunk.metadata.get("section_path") or "")
                    sections.append(
                        {
                            "path": [part for part in section_path.split(" / ") if part],
                            "paragraphs": [chunk.text],
                        }
                    )
                tree["sections"] = sections
                summary["tree"] = tree
            parsed_tables = parsed.get("tables")
            if isinstance(parsed_tables, Sequence) and parsed_tables:
                merged_tables, continuation_facts = merge_continuation_tables(
                    tuple(table for table in parsed_tables if isinstance(table, Mapping))
                )
                table_chunks, table_summary, table_degradations = self._markdown_table_chunks(
                    request,
                    merged_tables,
                    content_manifest_hash,
                    continuation_facts=tuple(continuation_facts),
                )
                chunks = [*chunks, *table_chunks]
                summary["chunk_count"] = len(chunks)
                summary["table_count"] = int(summary.get("table_count", 0)) + int(
                    table_summary["table_count"]
                )
                summary["row_groups"] = [
                    *list(summary.get("row_groups") or []),
                    *list(table_summary["row_groups"]),
                ]
                summary["continuation_tables"] = table_summary["continuation_tables"]
                degradations = (*degradations, *table_degradations)
            if mineru_probe_failed:
                sample = OCRSamplePlan.for_page_count(page_count or 1)
                local_usage_facts.append(
                    {
                        "kind": "local_usage",
                        "stage": "mineru_pipeline_probe",
                        "resource_kind": "document",
                        "item_count": len(sample.pages),
                        "page_count": page_count or 1,
                    }
                )
                summary["ocr"] = {
                    **dict(summary.get("ocr", {})),
                    "status": "degraded",
                    "probe": "provider_unavailable",
                    "sample_pages": list(OCRSamplePlan.for_page_count(page_count or 1).pages),
                }
                summary["contextual_input_ready"] = False
                degradations = (
                    *degradations,
                    *(
                        {
                            "kind": "contextual_retrieval_degraded",
                            "reason": "ocr_provider_unavailable",
                            "chunk_id": chunk.chunk_id,
                            "provider": "mineru-pipeline",
                        }
                        for chunk in chunks
                        if contextual_target(chunk)
                    ),
                )
        elif kind.startswith("image/") or kind in {"image", "图片"}:
            route_adapter = "image-vlm"
            context = {
                "media_kind": media_kind,
                "decorative": decorative_image,
                **dict(image_context or {}),
                "ocr_text": image_ocr_text or "",
            }
            dimensions = image_dimensions(raw)
            if dimensions is not None and not context.get("decorative"):
                page_area = context.get("page_area")
                context["decorative"] = is_decorative(
                    dimensions[0],
                    dimensions[1],
                    page_area=int(page_area) if isinstance(page_area, int) else None,
                )
                context["image_dimensions"] = {"width": dimensions[0], "height": dimensions[1]}
            if not context["ocr_text"] and self._image_ocr is not None:
                context["ocr_text"] = self._image_ocr(raw, context)
            chunks, summary, degradations = self._image_chunks(
                request,
                raw,
                content_manifest_hash,
                context=context,
            )
            image_provider_summary = summary.get("image_provider")
            if (
                isinstance(image_provider_summary, Mapping)
                and image_provider_summary.get("provider") == "internvl"
            ):
                local_usage_facts.append(
                    {
                        "kind": "local_usage",
                        "stage": "image_vlm_internvl",
                        "resource_kind": "image",
                        "item_count": 1,
                        "page_count": 1,
                    }
                )
        else:
            raise PlatformError("unsupported_media_type", "Media type is not supported", {}, 415)
        summary = dict(summary)
        if (
            route_adapter == "structured-table"
            or kind in {"text/csv", "csv"}
            or kind.endswith("csv")
        ):
            summary["metering"] = {"class": "table"}
        elif route_adapter == "image-vlm":
            provider_summary = summary.get("image_provider")
            summary["metering"] = {
                "class": "image",
                "image_provider": (
                    str(provider_summary.get("provider"))
                    if isinstance(provider_summary, Mapping)
                    and provider_summary.get("provider") is not None
                    else "none"
                ),
            }
        elif route_adapter in {"text", "mineru-pipeline"} and "metering" not in summary:
            summary["metering"] = {"class": "document"}
        chunks, summary, degradations = self._apply_contextual_retrieval(
            request,
            chunks,
            summary,
            degradations,
            document_text=document_text,
            media_kind=media_kind,
        )
        summary = dict(summary)
        summary["processing_list"] = self._processing_list(request, chunks)
        summary["media_kind"] = media_kind
        summary["document_profile"] = document_profile.to_mapping()
        summary["route"] = {
            "adapter": route_adapter,
            "media_kind": kind,
            "mineru": route_adapter == "mineru-pipeline",
            "vlm": route_adapter == "image-vlm",
            "structured_loader": route_adapter == "structured-table",
        }
        summary["processing_profile_version"] = profile
        summary.setdefault("page_count", page_count)
        summary.setdefault(
            "has_text_layer",
            has_text_layer if has_text_layer is not None else kind not in {"image", "图片"},
        )
        summary.setdefault("sheet_manifest", [])
        summary.setdefault("row_groups", [])
        summary.setdefault("failure_reason", None)
        if page_count:
            sample = OCRSamplePlan.for_page_count(page_count)
            summary.setdefault("ocr", {})
            summary["ocr"] = {**dict(summary["ocr"]), "sample_pages": list(sample.pages)}
        stage_resources = tuple(
            {
                "backend_kind": "index_chunk",
                "resource_id": f"{request.attempt_id}:{request.publication_id}:{chunk.chunk_id}",
                "attempt_id": request.attempt_id,
                "publication_id": request.publication_id,
                "fencing_token": request.fencing_token,
                "document_id": request.document_id,
                "document_version_id": request.document_version_id,
                "generation_id": request.expected_generation_id,
                "space_id": request.space_id,
            }
            for chunk in chunks
        )
        configured_models = {
            name: str(request.processing_config_snapshot[name])
            for name in (
                "parser",
                "ocr_model",
                "vlm_model",
                "embedding_model",
                "reranker_model",
                "cr_model",
            )
            if request.processing_config_snapshot.get(name) is not None
        }
        configured_prompts = {
            name: str(request.processing_config_snapshot[name])
            for name in ("prompt", "ocr_prompt", "vlm_prompt", "cr_prompt")
            if request.processing_config_snapshot.get(name) is not None
        }
        if self._contextual_provider is not None:
            configured_models["cr"] = self._contextual_provider.model
            configured_prompts["cr_prefix"] = CONTEXTUAL_PROMPT_SCHEMA_VERSION
        # §2.3：重放序列的快照固化了主模型/prompt 版本，执行端优先用冻结值，
        # 保证同代 attempt 的回执身份一致（初始序列快照无该键，走注入值）。
        model_version = (
            str(request.processing_config_snapshot["model_version"])
            if request.processing_config_snapshot.get("model_version") is not None
            else model_version
        )
        prompt_version = (
            str(request.processing_config_snapshot["prompt_version"])
            if request.processing_config_snapshot.get("prompt_version") is not None
            else prompt_version
        )
        receipt = IndexProcessingReceipt(
            job_id=request.job_id,
            attempt_id=request.attempt_id,
            fencing_token=request.fencing_token,
            publication_id=request.publication_id,
            document_id=request.document_id,
            document_version_id=request.document_version_id,
            input_content_hash=_sha256(raw),
            stage_resources=stage_resources,
            processing_config_version=profile,
            generation_id=request.expected_generation_id,
            authorization_fence=request.authorization_fence,
            model_version=model_version,
            prompt_version=prompt_version,
            processing_summary=summary,
            locator_snippet_integrity={
                "locators_valid": all(bool(chunk.locator is not None) for chunk in chunks),
                "snippets_valid": all(chunk.snippet is not None for chunk in chunks),
                "chunk_manifest_hash": content_manifest_hash,
            },
            index_component_results={
                "dense": {"state": "staged"},
                "sparse": {"state": "staged"},
                "tree": {
                    "state": "staged" if summary.get("tree", {}).get("tree_indexed") else "disabled"
                },
                "public_graph": {"state": "disabled"},
            },
            content_manifest_id=content_manifest_id,
            content_manifest_hash=content_manifest_hash,
            failure=None,
            degradations=tuple(degradations),
            ocr_low_confidence=bool(summary.get("ocr", {}).get("low_confidence", False)),
            ocr_low_confidence_fact=summary.get("ocr", {}).get("fact"),
            input_manifest_hash=request.input_manifest_hash,
            processing_profile_version=request.processing_profile_version,
            stage_resource_ids=tuple(str(item["resource_id"]) for item in stage_resources),
            model_versions={"primary": model_version, **configured_models},
            prompt_versions={"primary": prompt_version, **configured_prompts},
            space_id=request.space_id,
            operation=request.operation,
            base_active_version_id=request.base_active_version_id,
            index_revision_at_start=request.index_revision_at_start,
            object_manifest_ref=request.object_manifest_ref,
            processing_config_snapshot=dict(request.processing_config_snapshot),
        )
        receipt.validate_against(request)
        self._submit_local_usage(request, local_usage_facts)
        return ProcessingOutput(tuple(chunks), receipt)

    def _apply_contextual_retrieval(
        self,
        request: IndexStagingRequest,
        chunks: list[IndexChunk],
        summary: Mapping[str, Any],
        degradations: tuple[Mapping[str, Any], ...],
        *,
        document_text: str,
        media_kind: str,
    ) -> tuple[list[IndexChunk], Mapping[str, Any], tuple[Mapping[str, Any], ...]]:
        eligible = [chunk for chunk in chunks if contextual_target(chunk)]
        if (
            self._contextual_provider is None
            or not eligible
            or summary.get("contextual_input_ready") is False
        ):
            return chunks, summary, degradations

        metadata = {
            "document_id": request.document_id,
            "document_version_id": request.document_version_id,
            "media_kind": media_kind,
            "object_manifest_ref": request.object_manifest_ref,
            "processing_profile_version": request.processing_profile_version,
        }
        document = ContextualDocument(
            instance_id=str(request.processing_config_snapshot.get("index_namespace") or "default"),
            space_id=request.space_id,
            document_id=request.document_id,
            document_version_id=request.document_version_id,
            generation_id=request.expected_generation_id,
            metadata=metadata,
            full_text=document_text,
        )
        service = ContextualRetrievalService(
            provider=self._contextual_provider,
            cache=self._contextual_cache,
            concurrency=self._contextual_concurrency,
            max_attempts=self._contextual_max_attempts,
            token_limit=self._contextual_prefix_token_limit,
            token_counter=(self._contextual_token_counter or approximate_token_count),
            tokenization_config_version=(
                (
                    "runtime-tokenizer-v1"
                    if self._contextual_token_counter is not None
                    else CONTEXTUAL_TOKENIZATION_VERSION
                )
                + f":limit={self._contextual_prefix_token_limit}"
            ),
            usage_sink=lambda fact: self._submit_contextual_provider_usage(request, fact),
        )
        enhancement = service.enhance(document, chunks)
        enhanced_chunks = [
            replace(
                chunk,
                embedding_text=(
                    f"CONTEXTUAL RETRIEVAL\n{enhancement.contexts[chunk.chunk_id]}\n\n"
                    f"ORIGINAL CHUNK\n{chunk.embedding_text}"
                    if chunk.chunk_id in enhancement.contexts
                    else chunk.embedding_text
                ),
                metadata={
                    **dict(chunk.metadata),
                    "contextual_retrieval": (
                        "applied" if chunk.chunk_id in enhancement.contexts else "raw-fallback"
                    ),
                },
            )
            for chunk in chunks
        ]
        unit_summaries = [
            {
                "unit_id": unit.unit_id,
                "grouping": unit.grouping,
                "chunk_ids": [item.chunk_id for item in unit.chunks],
                "warmup_chunk_ids": list(unit.warmup_chunk_ids),
                "concurrent_chunk_ids": list(unit.concurrent_chunk_ids),
                "cache_mode": unit.cache_mode,
                "cache_outage": unit.cache_outage,
                "cache_reason": unit.cache_reason,
                "prefix_truncated": unit.prefix_truncated,
                "estimated_prefix_tokens": unit.estimated_prefix_tokens,
                "prompt_cache_hit_tokens": unit.prompt_cache_hit_tokens,
                "prompt_cache_miss_tokens": unit.prompt_cache_miss_tokens,
            }
            for unit in enhancement.units
        ]
        cr_summary = {
            "eligible_count": len(eligible),
            "enhanced_count": len(enhancement.contexts),
            "applied": bool(enhancement.contexts),
            "unit": "leaf_chunk",
            "model": self._contextual_provider.model,
            "model_revision": self._contextual_provider.model_revision,
            "provider": self._contextual_provider.provider,
            "prefix_units": unit_summaries,
            "prompt_cache_hit_tokens": sum(
                unit.prompt_cache_hit_tokens for unit in enhancement.units
            ),
            "prompt_cache_miss_tokens": sum(
                unit.prompt_cache_miss_tokens for unit in enhancement.units
            ),
        }
        result_summary = {**dict(summary), "cr": cr_summary}
        tree = dict(result_summary.get("tree") or {})
        if tree.get("tree_indexed"):
            tree["node_summaries"] = [
                {
                    "path": path,
                    "summary": "\n".join(
                        enhancement.contexts[chunk.chunk_id]
                        for chunk in group
                        if chunk.chunk_id in enhancement.contexts
                    ),
                    "model": self._contextual_provider.model,
                    "source": "contextual_retrieval",
                }
                for path, group in _group_chunks_by_tree_parent(enhanced_chunks)
            ]
            tree["node_summary_model"] = self._contextual_provider.model
            result_summary["tree"] = tree
        return (
            enhanced_chunks,
            result_summary,
            (
                *degradations,
                *enhancement.degradations,
            ),
        )

    @staticmethod
    def _processing_list(
        request: IndexStagingRequest, chunks: Sequence[IndexChunk]
    ) -> Mapping[str, Any]:
        return {
            "processing_list_id": f"processing_list:{request.publication_id}:{request.attempt_id}",
            "frozen": True,
            "items": [
                {
                    "chunk_id": chunk.chunk_id,
                    "contextual_retrieval": contextual_target(chunk),
                    "media_kind": chunk.media_kind,
                }
                for chunk in chunks
            ],
        }

    def _submit_contextual_provider_usage(
        self, request: IndexStagingRequest, fact: ContextualUsageFact
    ) -> None:
        submission = self._usage_submission
        prepare = getattr(submission, "prepare_provider_call", None)
        dispatch = getattr(submission, "mark_dispatching", None)
        complete = getattr(submission, "complete_provider_call", None)
        if submission is None or not callable(prepare) or not callable(complete):
            return
        ownership_values = request.usage_ownership
        if not isinstance(ownership_values, Mapping):
            return
        deadline = request.usage_deadline_at_utc or datetime.now(UTC) + timedelta(minutes=5)
        call_id = prepare(
            provider=fact.provider,
            model=fact.model,
            operation=fact.operation,
            execution_kind="document_processing",
            execution_id=request.attempt_id,
            attempt_id=request.attempt_id,
            generation_id=request.expected_generation_id,
            resource_id=f"{request.publication_id}:{fact.unit_id}:{fact.chunk_id}",
            deadline_utc=deadline,
            request_fingerprint=fact.request_fingerprint,
            replay_generation=request.usage_replay_generation,
        )
        if callable(dispatch):
            dispatch(call_id, started_at_provider=lambda: datetime.now(UTC))
        values = dict(ownership_values)
        values["source_space_ids"] = tuple(values.get("source_space_ids") or ())
        ownership = OwnershipSnapshot(**values)
        measurement = ProviderMeasurement(
            input_tokens=fact.input_tokens,
            prompt_cache_hit_tokens=fact.prompt_cache_hit_tokens,
            prompt_cache_miss_tokens=fact.prompt_cache_miss_tokens,
            output_tokens=fact.output_tokens,
            reasoning_tokens=None,
            image_count=None,
            visual_input_tokens=None,
            embedding_input_tokens=None,
            vector_count=None,
            measurement_sources={
                field: "provider_reported"
                for field in (
                    "input_tokens",
                    "prompt_cache_hit_tokens",
                    "prompt_cache_miss_tokens",
                    "output_tokens",
                )
                if getattr(fact, field) is not None
            },
        )
        complete(
            provider_call_id=call_id,
            measurement=measurement,
            ownership=ownership,
            result=fact.result,
            provider_request_id=fact.provider_request_id,
            started_at_utc=None,
        )

    def _submit_local_usage(
        self,
        request: IndexStagingRequest,
        facts: Sequence[Mapping[str, Any]],
    ) -> None:
        """Submit aggregated local expensive-stage usage through the typed port."""

        if self._usage_submission is None or not facts:
            return
        ownership_values = request.usage_ownership
        if not isinstance(ownership_values, Mapping):
            return
        values = dict(ownership_values)
        values["source_space_ids"] = tuple(values.get("source_space_ids") or ())
        ownership = OwnershipSnapshot(**values)
        meter_fields = (
            "item_count",
            "page_count",
            "input_bytes",
            "gpu_milliseconds",
            "cpu_milliseconds",
            "peak_vram_bytes",
        )
        started = datetime.now(UTC)
        for fact in facts:
            measurement = LocalMeasurement(
                **{field: fact.get(field) for field in meter_fields},
                measurement_sources={
                    field: "estimated" for field in meter_fields if fact.get(field) is not None
                },
            )
            self._usage_submission.submit_local_usage(
                execution_kind="document_processing",
                # 异步处理四元组的 execution_id 是 attempt_id（设计 §2.4.2）。
                execution_id=request.attempt_id,
                stage=str(fact.get("stage", "document_ingestion")),
                resource_kind=str(fact.get("resource_kind", "document")),
                measurement=measurement,
                ownership=ownership,
                result="succeeded",
                started_at_utc=started,
                replay_generation=request.usage_replay_generation,
            )

    def _text_chunks(
        self,
        request: IndexStagingRequest,
        text: str,
        media_kind: str,
        manifest_hash: str,
    ) -> tuple[list[IndexChunk], Mapping[str, Any], tuple[Mapping[str, Any], ...]]:
        sections = _sections(text)
        tree_indexed = bool(sections and any(path for path, _ in sections))
        if not sections and text.strip():
            sections = [("", text.strip())]
        chunks: list[IndexChunk] = []
        for section_index, (path, section) in enumerate(sections, start=1):
            block_index = 0
            # 段落锚点（B9-locator）：同遍产出——先按空行切出原始段落，再对
            # 超长段落做尺寸切块；同一段落切出的续块共享同一 paragraph。
            # 单 chunk section 不加锚（既有 locator 形状不变）。
            section_paragraphs = [p for p in re.split(r"\n\s*\n", section) if p.strip()]
            section_bodies: list[str] = []
            paragraph_of_body: list[int] = []
            for paragraph_number, paragraph in enumerate(section_paragraphs, start=1):
                pieces = _split_blocks_preserving_tables(
                    paragraph, maximum=self._text_chunk_max_chars
                )
                section_bodies.extend(pieces)
                paragraph_of_body.extend([paragraph_number] * len(pieces))
            for body, paragraph_number in zip(section_bodies, paragraph_of_body, strict=True):
                block_index += 1
                index = len(chunks) + 1
                compressed = self._compressor.compress(body, context={"section_path": path})
                locator: dict[str, Any] = {"section_path": path} if path else {}
                if len(section_bodies) > 1:
                    locator["paragraph"] = paragraph_number
                metadata: dict[str, Any] = {
                    "section_path": path,
                    "cr_parent_group": f"{section_index}:{block_index}",
                    "cr_unit": "chunk",
                }
                reverse_links = parse_reverse_links(body)
                if reverse_links:
                    metadata["reverse_links"] = [dict(link) for link in reverse_links]
                chunks.append(
                    IndexChunk(
                        chunk_id=f"chunk_{index}",
                        generation_id=request.expected_generation_id,
                        publication_id=request.publication_id,
                        document_id=request.document_id,
                        document_version_id=request.document_version_id,
                        space_id=request.space_id,
                        text=body,
                        embedding_text=compressed,
                        sparse_text=(f"{path}\n{body}" if path else body),
                        locator=locator,
                        snippet=body[:500],
                        media_kind=media_kind,
                        manifest_hash=manifest_hash,
                        metadata=metadata,
                    )
                )
        return (
            chunks,
            {
                "chunk_count": len(chunks),
                "page_count": 1,
                "image_count": 0,
                "table_count": 0,
                "ocr": {"sample_strategy": "head_middle_tail", "low_confidence": False},
                "tree": {
                    "tree_indexed": tree_indexed,
                    "tree_reason": "structure_signal" if tree_indexed else "no_structure",
                    "section_count": len(sections),
                },
                "cr": {"applied": bool(chunks), "unit": "chunk"},
            },
            (),
        )

    def _csv_chunks(
        self,
        request: IndexStagingRequest,
        text: str,
        manifest_hash: str,
    ) -> tuple[list[IndexChunk], Mapping[str, Any], tuple[Mapping[str, Any], ...]]:
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        if not rows or not rows[0]:
            raise PlatformError("table_parse_failed", "CSV has no header", {}, 422)
        headers = [str(item).strip() for item in rows[0]]
        data = rows[1:]
        ranges = _split_rows(data, len(headers))
        chunks: list[IndexChunk] = []
        row_groups: list[Mapping[str, Any]] = []
        for start, end in ranges:
            values = data[start:end]
            lines = [
                " | ".join(
                    f"{header}: {row[col] if col < len(row) else ''}"
                    for col, header in enumerate(headers)
                )
                for row in values
            ]
            # 行组超长按整行分组再切（A60）：单行内容不再被按字符截断；每个
            # chunk 的行号区间跟随该部分实际包含的行。
            line_index = 0
            for part in _split_lines_preserving_rows(lines, maximum=self._text_chunk_max_chars):
                part_first = line_index
                part_last = line_index + len(part) - 1
                line_index += len(part)
                row_start = start + part_first + 2
                row_end = start + part_last + 2
                index = len(chunks) + 1
                chunks.append(
                    IndexChunk(
                        chunk_id=f"chunk_{index}",
                        generation_id=request.expected_generation_id,
                        publication_id=request.publication_id,
                        document_id=request.document_id,
                        document_version_id=request.document_version_id,
                        space_id=request.space_id,
                        text="\n".join(part),
                        embedding_text="\n".join(part),
                        sparse_text="\n".join(part),
                        locator={
                            "sheet": "CSV",
                            "a1_range": (f"A{row_start}:{_column_name(len(headers))}{row_end}"),
                        },
                        snippet=None,
                        media_kind="text/csv",
                        manifest_hash=manifest_hash,
                        metadata={
                            "headers": headers,
                            "row_start": row_start,
                            "row_end": row_end,
                            "table": True,
                            "block": index,
                        },
                    )
                )
                row_groups.append(
                    {"sheet": "CSV", "start": row_start, "end": row_end, "block": index}
                )
        total_blocks = len(chunks)
        chunks = [
            replace(chunk, metadata={**chunk.metadata, "total_blocks": total_blocks})
            for chunk in chunks
        ]
        return (
            chunks,
            {
                "chunk_count": len(chunks),
                "page_count": 1,
                "image_count": 0,
                "table_count": len(chunks),
                "ocr": {},
                "tree": {"tree_indexed": False, "tree_reason": "table"},
                "cr": {"applied": False, "unit": "table_header"},
                "sheet_count": 1,
                "headers": headers,
                "sheet_manifest": [{"sheet": "CSV", "headers": headers, "row_count": len(rows)}],
                "row_groups": row_groups,
            },
            (),
        )

    def _markdown_table_chunks(
        self,
        request: IndexStagingRequest,
        tables: Sequence[Mapping[str, Any]],
        manifest_hash: str,
        *,
        continuation_facts: Sequence[Mapping[str, Any]] = (),
    ) -> tuple[list[IndexChunk], Mapping[str, Any], tuple[Mapping[str, Any], ...]]:
        """Emit complete markdown tables; continuation tables are merged already."""

        chunks: list[IndexChunk] = []
        row_groups: list[Mapping[str, Any]] = []
        degradations: list[Mapping[str, Any]] = []
        planned: list[tuple[Mapping[str, Any], tuple[int, int]]] = []
        for table in tables:
            headers = [str(item) for item in (table.get("headers") or ())]
            rows = [[str(cell) for cell in row] for row in (table.get("rows") or ())]
            if not headers:
                degradations.append({"code": "table_parse_failed", "reason": "missing_headers"})
                continue
            for start, end in _split_rows(rows, len(headers)):
                planned.append((table, (start, end)))
        total_blocks = len(planned)
        for block, (table, (start, end)) in enumerate(planned, start=1):
            headers = [str(item) for item in (table.get("headers") or ())]
            rows = [[str(cell) for cell in row] for row in (table.get("rows") or ())]
            group = rows[start:end]
            markdown = _markdown_table(headers, group)
            key_values = _key_value_rows(headers, group)
            page_start = int(table.get("page_start", 1) or 1)
            page_end = int(table.get("page_end", page_start) or page_start)
            context_title = str(table.get("title", "") or "")
            index = len(chunks) + 1
            chunks.append(
                IndexChunk(
                    chunk_id=f"chunk_{index}",
                    generation_id=request.expected_generation_id,
                    publication_id=request.publication_id,
                    document_id=request.document_id,
                    document_version_id=request.document_version_id,
                    space_id=request.space_id,
                    text=markdown,
                    embedding_text=(
                        f"{context_title}\n{key_values}" if context_title else key_values
                    ),
                    sparse_text=key_values,
                    locator=(
                        {"page": page_start, "page_end": page_end}
                        if page_start != page_end
                        else {"page": page_start}
                    ),
                    snippet=markdown[:500],
                    media_kind="table/markdown",
                    manifest_hash=manifest_hash,
                    metadata={
                        "headers": headers,
                        "table": True,
                        "table_title": context_title,
                        "block": block,
                        "total_blocks": total_blocks,
                        "page_start": page_start,
                        "page_end": page_end,
                        "row_start": start + 1,
                        "row_end": end,
                        "continuation": bool(
                            table.get("page_start") is not None
                            and int(table.get("page_start", 1)) != page_end
                        ),
                    },
                )
            )
            row_groups.append(
                {
                    "sheet": f"page:{page_start}-{page_end}",
                    "start": start + 1,
                    "end": end,
                    "block": block,
                    "total_blocks": total_blocks,
                }
            )
        return (
            chunks,
            {
                "table_count": len(tables),
                "row_groups": row_groups,
                "continuation_tables": [dict(fact) for fact in continuation_facts],
            },
            tuple(degradations),
        )

    def _xlsx_chunks(
        self,
        request: IndexStagingRequest,
        content: bytes,
        manifest_hash: str,
    ) -> tuple[list[IndexChunk], Mapping[str, Any], tuple[Mapping[str, Any], ...]]:
        try:
            from openpyxl import load_workbook  # type: ignore[import-untyped]

            merged_ranges_by_sheet = _xlsx_merged_ranges(
                content, maximum_cells=self._xlsx_merged_cells_max
            )
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except PlatformError:
            raise
        except Exception as exc:
            raise PlatformError(
                "table_parse_failed", "Excel workbook could not be parsed", {}, 422
            ) from exc
        chunks: list[IndexChunk] = []
        table_count = 0
        sheet_count = len(workbook.sheetnames)
        sheet_manifest: list[Mapping[str, Any]] = []
        row_groups: list[Mapping[str, Any]] = []
        try:
            from openpyxl.utils.cell import range_boundaries  # type: ignore[import-untyped]

            for worksheet in workbook.worksheets:
                merged_ranges = list(merged_ranges_by_sheet.get(worksheet.title, ()))
                merged_cells = _MergedCellResolver(merged_ranges)
                origin_values: dict[tuple[int, int], str] = {}
                collected: list[tuple[int, list[str]]] = []
                row_count = 0
                for row_number, raw_row in enumerate(
                    worksheet.iter_rows(values_only=True), start=1
                ):
                    row_count = row_number
                    merged_cells.advance(row_number)
                    row: list[str] = []
                    for column, value in enumerate(raw_row, start=1):
                        origin = merged_cells.origin_at(column)
                        rendered = "" if value is None else str(value)
                        if origin == (row_number, column):
                            origin_values[origin] = rendered
                        elif origin is not None:
                            rendered = origin_values.get(origin, rendered)
                        row.append(rendered)
                    if any(cell.strip() for cell in row):
                        collected.append((row_number, row))
                if not collected:
                    raise PlatformError("table_parse_failed", "Excel sheet has no header", {}, 422)
                header_start = collected[0][0]
                header_depth = 1
                for reference in merged_ranges:
                    min_column, min_row, _max_column, max_row = range_boundaries(reference)
                    del min_column
                    if min_row < max_row and min_row <= header_start + 2:
                        header_depth = max(header_depth, max_row - header_start + 1)
                by_row = dict(collected)
                header_rows = [
                    by_row.get(row_number, [])
                    for row_number in range(header_start, header_start + header_depth)
                ]
                headers = _flatten_multilevel_header(header_rows)
                data_rows = [
                    (row_number, row)
                    for row_number, row in collected
                    if row_number >= header_start + header_depth
                ]
                sheet_manifest.append(
                    {
                        "sheet": worksheet.title,
                        "headers": headers,
                        "merged_ranges": merged_ranges,
                        "row_count": row_count,
                    }
                )
                width = len(headers)
                for start, end in _split_rows([row for _, row in data_rows], width):
                    group = data_rows[start:end]
                    if not group:
                        continue
                    start_row = group[0][0]
                    end_row = group[-1][0]
                    total_rows = [
                        row_number
                        for row_number, _ in group
                        if row_number in merged_cells.horizontally_merged_rows
                    ]
                    body = _key_value_rows(headers, [row for _, row in group])
                    if not body.strip():
                        continue
                    table_count += 1
                    block_number = len(row_groups) + 1
                    row_groups.append(
                        {
                            "sheet": worksheet.title,
                            "start": start_row,
                            "end": end_row,
                            "block": block_number,
                            "total_rows": total_rows,
                        }
                    )
                    chunks.append(
                        IndexChunk(
                            chunk_id=f"chunk_{len(chunks) + 1}",
                            generation_id=request.expected_generation_id,
                            publication_id=request.publication_id,
                            document_id=request.document_id,
                            document_version_id=request.document_version_id,
                            space_id=request.space_id,
                            text=body,
                            embedding_text=body,
                            sparse_text=body,
                            locator={
                                "sheet": worksheet.title,
                                "a1_range": f"A{start_row}:{_column_name(width)}{end_row}",
                            },
                            snippet=None,
                            media_kind="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            manifest_hash=manifest_hash,
                            metadata={
                                "headers": headers,
                                "row_start": start_row,
                                "row_end": end_row,
                                "table": True,
                                "merged_ranges": merged_ranges,
                                "total_rows": total_rows,
                                "block": block_number,
                            },
                        )
                    )
        finally:
            workbook.close()
        total_blocks = len(row_groups)
        chunks = [
            replace(chunk, metadata={**chunk.metadata, "total_blocks": total_blocks})
            for chunk in chunks
        ]
        return (
            chunks,
            {
                "chunk_count": len(chunks),
                "page_count": 0,
                "image_count": 0,
                "table_count": table_count,
                "sheet_count": sheet_count,
                "ocr": {},
                "tree": {"tree_indexed": False, "tree_reason": "table"},
                "cr": {"applied": False, "unit": "table_header"},
                "sheet_manifest": sheet_manifest,
                "row_groups": row_groups,
            },
            (),
        )

    def _structured_chunks(
        self,
        request: IndexStagingRequest,
        text: str,
        kind: str,
        manifest_hash: str,
    ) -> tuple[list[IndexChunk], Mapping[str, Any], tuple[Mapping[str, Any], ...]]:
        try:
            if kind in {"application/json", "json"}:
                value = json.loads(text)
            elif kind in {"application/toml", "toml"}:
                value = tomllib.loads(text)
            elif kind in {"application/xml", "text/xml", "xml"}:
                if re.search(r"<!\s*ENTITY\b", text, flags=re.IGNORECASE):
                    raise PlatformError(
                        "structured_parse_failed",
                        "XML entity declarations are not supported",
                        {},
                        422,
                    )
                root = ElementTree.fromstring(text)
                # 保留层级（A61）：子树递归转嵌套映射/列表，同名兄弟折叠为列表，
                # 由统一的深嵌套/扁平化判据分流。
                value = {root.tag: _element_to_value(root)}
            else:
                value = _parse_yaml_subset(text)
        except (
            ValueError,
            TypeError,
            json.JSONDecodeError,
            tomllib.TOMLDecodeError,
            ElementTree.ParseError,
        ) as exc:
            raise PlatformError(
                "structured_parse_failed", "Structured input could not be parsed", {}, 422
            ) from exc
        if isinstance(value, list) and value and all(isinstance(item, Mapping) for item in value):
            headers = sorted({str(key) for item in value for key in item})
            rows = [[str(item.get(header, "")) for header in headers] for item in value]
            chunks, summary, degradations = self._table_chunks(
                request,
                headers,
                rows,
                manifest_hash,
                sheet="structured",
                media_kind=kind,
            )
            summary = dict(summary)
            summary["metering"] = {"class": "table"}
            return chunks, summary, degradations
        token_count = approximate_token_count(text)
        # A61 分流判据：schema 特征组合（$schema/schema/properties/required 等）或
        # 深嵌套（无 schema 键）走代码/配置路径——单独的顶层 "type" 不再误入，
        # 深嵌套对象不再键值对扁平化。
        schema_like = isinstance(value, Mapping) and any(
            key in value for key in _STRUCTURED_SCHEMA_KEYS
        )
        deeply_nested = (
            isinstance(value, Mapping)
            and _nesting_depth(value) >= STRUCTURED_NESTING_DEPTH_THRESHOLD
        )
        if schema_like or deeply_nested:
            chunks, summary, degradations = self._code_chunks(
                request,
                text,
                manifest_hash,
                metering_class="config",
            )
            summary = dict(summary)
            summary["metering"] = {"class": "config", "token_count": token_count}
            if token_count <= 500:
                summary["contextual_input_ready"] = False
            return chunks, summary, degradations
        pairs = _flatten(value)
        body = "\n".join(f"{key}: {item}" for key, item in pairs)
        large = token_count > 500
        chunk = IndexChunk(
            chunk_id="chunk_1",
            generation_id=request.expected_generation_id,
            publication_id=request.publication_id,
            document_id=request.document_id,
            document_version_id=request.document_version_id,
            space_id=request.space_id,
            text=body or "{}",
            embedding_text=(
                self._compressor.compress(body, context={"kind": "structured"}) if large else body
            )
            or "{}",
            sparse_text=body or "{}",
            locator={},
            snippet=body[:500] or "{}",
            media_kind=kind,
            manifest_hash=manifest_hash,
            metadata={
                "keys": [key for key, _ in pairs],
                "cr_unit": "chunk" if large else "key_path",
            },
        )
        return (
            [chunk],
            {
                "chunk_count": 1,
                "page_count": 0,
                "image_count": 0,
                "table_count": 0,
                "ocr": {},
                "tree": {"tree_indexed": False, "tree_reason": "structured"},
                "cr": {"applied": large, "unit": "chunk" if large else "key_path"},
                "metering": {"class": "config", "token_count": token_count},
                **({} if large else {"contextual_input_ready": False}),
            },
            (),
        )

    def _table_chunks(
        self,
        request: IndexStagingRequest,
        headers: Sequence[str],
        rows: Sequence[Sequence[str]],
        manifest_hash: str,
        *,
        sheet: str,
        media_kind: str,
    ) -> tuple[list[IndexChunk], Mapping[str, Any], tuple[Mapping[str, Any], ...]]:
        ranges = _split_rows(rows, len(headers))
        chunks: list[IndexChunk] = []
        for number, (start, end) in enumerate(ranges, start=1):
            body = "\n".join(
                " | ".join(
                    f"{header}: {row[column] if column < len(row) else ''}"
                    for column, header in enumerate(headers)
                )
                for row in rows[start:end]
            )
            chunks.append(
                IndexChunk(
                    chunk_id=f"chunk_{number}",
                    generation_id=request.expected_generation_id,
                    publication_id=request.publication_id,
                    document_id=request.document_id,
                    document_version_id=request.document_version_id,
                    space_id=request.space_id,
                    text=body,
                    embedding_text=body,
                    sparse_text=body,
                    locator={
                        "sheet": sheet,
                        "a1_range": f"A{start + 2}:{_column_name(len(headers))}{end + 1}",
                    },
                    snippet=None,
                    media_kind=media_kind,
                    manifest_hash=manifest_hash,
                    metadata={"headers": list(headers), "table": True},
                )
            )
        return (
            chunks,
            {
                "chunk_count": len(chunks),
                "page_count": 0,
                "image_count": 0,
                "table_count": len(chunks),
                "ocr": {},
                "tree": {"tree_indexed": False, "tree_reason": "structured_table"},
                "cr": {"applied": False, "unit": "table_header"},
                "sheet_manifest": [{"sheet": sheet, "headers": list(headers)}],
                "row_groups": [
                    {"sheet": sheet, "start": start + 2, "end": end + 1} for start, end in ranges
                ],
            },
            (),
        )

    def _code_chunks(
        self,
        request: IndexStagingRequest,
        text: str,
        manifest_hash: str,
        *,
        metering_class: str = "code",
    ) -> tuple[list[IndexChunk], Mapping[str, Any], tuple[Mapping[str, Any], ...]]:
        symbols: list[tuple[str, str]] = []
        module_level: list[str] = []
        python_tree: ast.Module | None = None
        try:
            python_tree = ast.parse(text)
        except SyntaxError:
            python_tree = None
        if python_tree is not None:
            lines = text.splitlines()
            for node in python_tree.body:
                if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef, ast.ClassDef)):
                    # 装饰器行属于符号本身（lineno 指向 def/class）。
                    start_line = node.lineno
                    for decorator in getattr(node, "decorator_list", ()):
                        start_line = min(start_line, decorator.lineno)
                    symbols.append((node.name, "\n".join(lines[start_line - 1 : node.end_lineno])))
                else:
                    # import/模块级赋值等非符号语句落入模块级 chunk（A61），
                    # 不随符号切分丢弃。
                    module_level.extend(lines[node.lineno - 1 : node.end_lineno])
        if python_tree is None and not symbols and ("function " in text or "class " in text):
            matches = list(
                re.finditer(
                    r"^(?:export\s+)?(?:async\s+)?(?:function|class)\s+([A-Za-z_$][\w$]*)",
                    text,
                    re.MULTILINE,
                )
            )
            for index, match in enumerate(matches):
                end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
                symbols.append((match.group(1), text[match.start() : end].strip()))
        module_text = "\n".join(module_level).strip()
        # 模块级 chunk 在前（import/常量通常位于文件头），符号 chunk 保持顺序。
        ordered: list[tuple[str, str]] = (
            [("module", module_text)] if module_text else []
        ) + symbols
        if not ordered:
            ordered = [("module", text.strip())]
        chunks: list[IndexChunk] = []
        for name, symbol in ordered:
            if not symbol.strip():
                continue
            for body in _split_text(symbol, maximum=self._text_chunk_max_chars):
                index = len(chunks) + 1
                chunks.append(
                    IndexChunk(
                        chunk_id=f"chunk_{index}",
                        generation_id=request.expected_generation_id,
                        publication_id=request.publication_id,
                        document_id=request.document_id,
                        document_version_id=request.document_version_id,
                        space_id=request.space_id,
                        text=body,
                        embedding_text=self._compressor.compress(body, context={"symbol": name}),
                        # 共享符号名进 BM25 字段（08-5-14）：长符号被切分后，
                        # 续块不再包含 def/类声明行；符号名前置保证符号名查询
                        # 在任意切片上都可命中。文本路径 sparse 字段不变。
                        sparse_text=f"{name}\n{body}",
                        locator={"section_path": name},
                        snippet=body[:500],
                        media_kind="code",
                        manifest_hash=manifest_hash,
                        metadata={"symbol": name, "cr_unit": "symbol"},
                    )
                )
        return (
            chunks,
            {
                "chunk_count": len(chunks),
                "page_count": 0,
                "image_count": 0,
                "table_count": 0,
                "ocr": {},
                "tree": {"tree_indexed": False, "tree_reason": "code"},
                "cr": {"applied": True, "unit": "symbol"},
                "metering": {
                    "class": metering_class,
                    "token_count": approximate_token_count(text),
                },
            },
            (),
        )

    def _image_chunks(
        self,
        request: IndexStagingRequest,
        content: bytes,
        manifest_hash: str,
        *,
        context: Mapping[str, Any],
    ) -> tuple[list[IndexChunk], Mapping[str, Any], tuple[Mapping[str, Any], ...]]:
        if context.get("decorative"):
            return (
                [],
                {
                    "chunk_count": 0,
                    "page_count": 0,
                    "image_count": 1,
                    "table_count": 0,
                    "ocr": {"applied": False},
                    "tree": {"tree_indexed": False, "tree_reason": "decorative_image"},
                    "cr": {"applied": False, "unit": "image"},
                    "filtered_image_count": 1,
                },
                ({"code": "image_not_indexable", "reason": "decorative"},),
            )
        result = (
            normalize_description(self._image_describer(content, context))
            if self._image_describer is not None
            else None
        )
        description = result.text.strip() if result is not None else ""
        caption = str(context.get("caption", "")).strip()
        ocr_text = str(context.get("ocr_text", "")).strip()
        index_text = "\n".join(part for part in (caption, description, ocr_text) if part)
        provider_summary = {
            "provider": result.provider if result is not None else "none",
            "degraded": bool(result.degraded) if result is not None else not bool(index_text),
            "reason": result.reason if result is not None else "no_provider",
            **(
                {"usage": dict(result.usage)}
                if result is not None and result.usage is not None
                else {}
            ),
        }
        if result is not None and result.degraded and result.reason == "degraded_no_text":
            return (
                [],
                {
                    "chunk_count": 0,
                    "page_count": 0,
                    "image_count": 1,
                    "table_count": 0,
                    "ocr": {"applied": bool(ocr_text)},
                    "tree": {"tree_indexed": False, "tree_reason": "image"},
                    "cr": {"applied": False, "unit": "image_description"},
                    "image_provider": provider_summary,
                },
                ({"code": "degraded_no_text", "reason": "vlm_disabled_no_text"},),
            )
        if not index_text:
            return (
                [],
                {
                    "chunk_count": 0,
                    "page_count": 0,
                    "image_count": 1,
                    "table_count": 0,
                    "ocr": {"applied": bool(ocr_text)},
                    "tree": {"tree_indexed": False, "tree_reason": "image"},
                    "cr": {"applied": False, "unit": "image_description"},
                    "image_provider": provider_summary,
                },
                ({"code": "image_not_indexable", "reason": "no_usable_text"},),
            )
        return (
            [
                IndexChunk(
                    chunk_id="chunk_1",
                    generation_id=request.expected_generation_id,
                    publication_id=request.publication_id,
                    document_id=request.document_id,
                    document_version_id=request.document_version_id,
                    space_id=request.space_id,
                    text=index_text,
                    embedding_text=index_text,
                    sparse_text=index_text,
                    locator={},
                    snippet=None,
                    media_kind="image",
                    manifest_hash=manifest_hash,
                    metadata={
                        "image_description": bool(description),
                        "image_provider": provider_summary,
                        "reverse_links": list(context.get("reverse_links", ())),
                    },
                )
            ],
            {
                "chunk_count": 1,
                "page_count": 0,
                "image_count": 1,
                "table_count": 0,
                "ocr": {"applied": bool(ocr_text)},
                "tree": {"tree_indexed": False, "tree_reason": "image"},
                "cr": {"applied": False, "unit": "image_description"},
                "image_provider": provider_summary,
            },
            (
                ({"code": "image_degraded", "reason": result.reason},)
                if result is not None and result.degraded
                else ()
            ),
        )


__all__ = [
    "ContentProcessor",
    "IdentityCompression",
    "OCRSamplePlan",
    "ProcessingOutput",
    "parse_reverse_links",
]
